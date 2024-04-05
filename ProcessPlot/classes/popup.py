"""
Copyright (c) 2021 Adam Solchenberger asolchenberger@gmail.com, Jason Engman engmanj@gmail.com

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

from ast import Pass
from enum import auto
from http.client import PARTIAL_CONTENT
from logging.config import valid_ident
from pkgutil import iter_modules
import numpy
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from urllib.parse import non_hierarchical
import gi, os, json, datetime, time

from pycomm3 import parse_connection_path
from numpy import maximum, nonzero
gi.require_version('Gtk', '3.0')
from gi.repository import Gtk, Gdk, GdkPixbuf, GObject
import re
from Public.widgets.checkbox import CheckBoxWidget

PUBLIC_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)),  'Public')


class BaseSettingsPopoup(Gtk.Dialog):

  def __init__(self, parent, title,app):
    super().__init__(title=title, transient_for = parent,flags=0) 
    self.app = app
    self.build_window()
    self.content_area = self.get_content_area()

    self.dialog_window = Gtk.Box(width_request=600,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window)
    ### -title bar- ####
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    ### -content area- ####
    self.base_area = Gtk.Box(spacing = 10,orientation=Gtk.Orientation.VERTICAL,margin = 20)
    self.scroll = Gtk.ScrolledWindow(width_request = 1400,height_request = 600)
    self.scroll.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
    self.scroll.add(self.base_area)
    self.dialog_window.pack_start(self.scroll,1,1,1)
    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)

    self.build_header(title)
    self.build_base()
    self.build_footer()
    self.show_all()
  
  def build_window(self, *args):
    self.set_default_size(1500, 700)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(False)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def add_style(self, item,style):
    sc = item.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def close_popup(self, button):
    self.destroy()

  def build_header(self,title):
    #header
    title = Gtk.Label(label=title)
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-18')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)
    self.pin_button = Gtk.Button(width_request = 20)
    self.pin_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.pin_button.add(image)
    self.title_bar.pack_end(self.pin_button,0,0,1)
    sc = self.pin_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self):
    pass

  def build_footer(self):
    pass

  def display_msg(self,msg,*args):
    popup = PopupMessage(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      return True
    else:
      return False

  def open_numpad(self,button,widget_obj,params,*args):
    numpad = ValueEnter(self,widget_obj,params)
    response = numpad.run()
    if response == Gtk.ResponseType.NO:
      pass
    else:
      pass
    numpad.destroy()

################################ change styling of connect toggle button , add connect to conx database
################################ When deleting a connection the row stays in the connection specific table
################################ Need to update connection toggle buttons on regular basis and when building page
################################ Add button on tags row to include on legend besides pens
################################ Add ability to save data to config file and upload data back to connection from file (ConfigUtility)
################################ 
################################



class PenSettingsPopup(Gtk.Dialog):

  def __init__(self, parent,app):
    super().__init__(transient_for = parent,flags=0) 
    self.app = app
    self.parent = parent
    self.tags_available = {}
    self.connections_available = {}
    self.connections_list = []
    self.pens_available = {}
    self.chart_filter = 'All'
    self.unsaved_changes_present = False
    self.unsaved_pen_rows = {}
    self.pen_column_names = ['id', 'chart_id', 'tag_id', 'connection_id', 'visible', 
                      'color', 'weight','scale_minimum','scale_maximum', 
                      'scale_lock', 'scale_auto']
    self.db_session = self.app.settings_db.session
    self.db_model = self.app.settings_db.models['pen']
    self.Tbl = self.db_model
    self.build_window()
    self.content_area = self.get_content_area()
    self.get_available_tags('c_id')
    self.get_available_connections()
    self.get_available_pens()

    self.dialog_window = Gtk.Box(width_request=600,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window)
    ### -title bar- ####
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    ### -content area- ####
    self.base_area = Gtk.Box(spacing = 10,orientation=Gtk.Orientation.VERTICAL,margin = 20)
    self.scroll = Gtk.ScrolledWindow(width_request = 1400,height_request = 600)
    self.scroll.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
    self.scroll.add(self.base_area)
    self.dialog_window.pack_start(self.scroll,1,1,1)
    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)

    self.build_header("Pen Settings")
    self.build_base()
    self.build_footer()
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(1400, 600)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(False)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self,title):
    #header
    db_c_num = 'All'
    selections = ['All',"Chart 1",'Chart 2','Chart 3',"Chart 4",'Chart 5','Chart 6',"Chart 7",'Chart 8','Chart 9',
                  "Chart 10",'Chart 11','Chart 12',"Chart 13",'Chart 14','Chart 15','Chart 16']
    self.c_num = Gtk.ComboBoxText()
    self.c_num.set_entry_text_column(0)
    for x in selections:
        self.c_num.append_text(x)
    try:
      idx = selections.index(str(db_c_num))
    except IndexError:
      idx = 0
    self.c_num.set_active(idx)
    sc = self.c_num.get_style_context()
    sc.add_class('ctrl-combo')
    self.c_num.connect("changed", self.filter_disp_chart)
    self.title_bar.pack_start(self.c_num,0,0,1)

    self.add_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/AddPen.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.add_button.add(image)
    sc = self.add_button.get_style_context()
    sc.add_class('ctrl-button')
    self.title_bar.pack_start(self.add_button,0,0,0)
    self.add_button.connect('clicked',self.create_new_pen)

    pen_import = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Import.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    pen_import.add(image)
    sc = pen_import.get_style_context()
    sc.add_class('ctrl-button')
    self.title_bar.pack_start(pen_import,0,0,0)
    pen_import.connect('clicked',self.fileChooser_open,"")

    pen_export = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Export.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    pen_export.add(image)
    sc = pen_export.get_style_context()
    sc.add_class('ctrl-button')
    self.title_bar.pack_start(pen_export,0,0,0)
    pen_export.connect('clicked',self.fileChooser_save,"")

    title = Gtk.Label(label=title,width_request = 1000)
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-18')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    self.pin_button = Gtk.Button(width_request = 20)
    self.pin_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.pin_button.add(image)
    self.title_bar.pack_end(self.pin_button,0,0,1)
    sc = self.pin_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self):
    self.pen_settings = []
    self.pen_row_num = 1
    #header = self.db_session.query(self.Tbl).first()
    #self.pen_column_names = header.__table__.columns
    self.pen_grid = Gtk.Grid(column_homogeneous=False,column_spacing=20,row_spacing=10)
    self.base_area.add(self.pen_grid)
    #header
    self.add_column_names()
    self.add_pen_rows(self.chart_filter)
    self.show_all()

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.save_settings,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    self.apply_button.connect('clicked',self.save_settings,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    #self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def filter_disp_chart(self,chart_filter,*args):
    temp = chart_filter.get_active_text()
    val = temp.strip('Chart ')
    self.chart_filter = val
    self.remove_pen_rows()
    self.add_column_names()
    self.add_pen_rows(self.chart_filter)

  def add_column_names(self,*args):
    labels = ['Chart','Enabled','Connection', 'Tag', 'Color',
      'Width', 'Scale Min', 'Scale Max', 'Auto Scale','Lock Scale','Save',''] # may want to create a table in the db for column names
    for l_idx in range(len(labels)):
        l = Gtk.Label(labels[l_idx])
        sc = l.get_style_context()
        sc.add_class('text-black-color')
        sc.add_class('font-14')
        sc.add_class('font-bold')
        self.pen_grid.attach(l, l_idx, 0, 1, 1)

  def create_delete_button(self,pen_id,row,*args):
    self.delete_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Delete.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.delete_button.add(image)
    sc = self.delete_button.get_style_context()
    sc.add_class('ctrl-button')
    self.pen_grid.attach(self.delete_button,11,row,1,1)
    self.delete_button.connect('clicked',self.confirm,pen_id)
  
  def delete_row(self,row_id,*args):
    settings = self.db_session.query(self.Tbl).filter(self.Tbl.id == row_id).first()
    self.db_session.query(self.Tbl).filter(self.Tbl.id == row_id).delete()
    self.db_session.commit()
    self.remove_pen_rows()
    self.add_column_names()
    self.get_available_pens()
    self.add_pen_rows(self.chart_filter)
    self.delete_pen_object(row_id,settings.chart_id)

  def add_pen_rows(self,chart_filter,*args):
    #pen row
    if not self.pens_available:
      self.pen_grid.set_column_homogeneous(True)
    else:
      self.pen_grid.set_column_homogeneous(False)   
    temp = {}
    for key, pen in self.pens_available.items():
      if chart_filter != 'All':
        if str(pen['chart_id']) == chart_filter:
          temp[key] = pen
      else:
        temp = self.pens_available
    for key, pen in temp.items():
        row = Pen_row(pen,self.pen_grid,self.pen_row_num,self.app,self)
        self.create_delete_button(pen['id'],self.pen_row_num)
        self.pen_row_num += 1
        self.pen_settings.append(row)
    self.show_all()

  def remove_pen_rows(self,*args):
    rows = self.pen_grid.get_children()
    for items in rows:
      self.pen_grid.remove(items)
  
  def create_new_pen(self,*args):
    if self.chart_filter == 'All':
      c_id = 1
    else:
      c_id = int(self.chart_filter)
    new = self.Tbl(chart_id = c_id)
    self.db_session.add(new)
    self.db_session.commit()    
    self.db_session.refresh(new)  #Retrieves newly created pen id from the database (new.id)
    self.insert_pen_row()
    self.create_pen_object(new.id,c_id)

  def create_pen_object(self,id,chart_id,*args):
    self.app.charts[chart_id].add_pen(id)
    self.get_available_pens()
  
  def delete_pen_object(self,id,chart_id,*args):
    self.app.charts[chart_id].delete_pen(id)
    self.get_available_pens()

  def insert_pen_row(self,*args):
    self.pen_grid.set_column_homogeneous(False) 
    self.pen_grid.insert_row(1)
    last_pen = self.db_session.query(self.Tbl).order_by(self.Tbl.id.desc()).first()
    params = {}
    for c in self.pen_column_names:
      params[c] = getattr(last_pen, c)
    row = Pen_row(params,self.pen_grid,1,self.app,self)
    self.create_delete_button(params['id'],1)
    params.clear()
    self.pen_row_num += 1
    self.pen_settings.append(row)
    self.show_all()

  def confirm(self, button,pen_id,msg="Are you sure you want to delete this pen?", args=[]):
    popup = PopupConfirm(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      self.delete_row(pen_id)
      return True
    else:
      return False

  def unsaved_changes(self,status,pen_row,id,*args):
    #This method holds references to pen rows for saveall changes on OK button
    self.unsaved_changes_present = status
    if status:
      self.unsaved_pen_rows[id] = pen_row
    else:
      del self.unsaved_pen_rows[id]

  def save_settings(self,*args):
    #when clicking ok button, save all unsaved settings, clear unsaved dictionary, and close popup
    for key,pen_row in self.unsaved_pen_rows.items():
      pen_row.update_db()
    self.unsaved_changes_present = False
    self.unsaved_pen_rows = {}
    self.close_popup(None)

  def get_available_tags(self,c_id,*args):
    new_params = {}
    count = 1
    self.conx_tags = {}
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      tag_items = conx_obj.return_tag_parameters()  #return list of tag parameters from the specific connection
      for tag_id,tag_obj in conx_obj.get('tags').items():
        for c in tag_items:
          new_params[c] = getattr(tag_obj, c)
        self.conx_tags[count] = new_params
        new_params = {}
        count += 1
      self.tags_available[conx_id]= self.conx_tags
      self.conx_tags = {}
      count = 1

  def get_available_connections(self,*args):
    conx_items = ['id', 'connection_type', 'description']
    new_params = {}
    count = 1
    self.connections_available = {0: {'id': '', 'connection_type': 0, 'description': ''}}
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      for c in conx_items:
        new_params[c] = getattr(conx_obj, c)
      self.connections_available[count] = new_params
      new_params = {}
      count += 1
    self.connections_list = []
    for key, conx in self.connections_available.items():
      self.connections_list.append(conx['id'])

  def get_available_pens(self,*args):
    params = {}
    pen = {}
    settings = self.db_session.query(self.Tbl).order_by(self.Tbl.id)
    for pen_param in settings:
      for c in self.pen_column_names:
        pen[c] = getattr(pen_param, c)
      params[pen['id']] = pen
      pen = {}
    self.pens_available = params

  def fileChooser_open(self,param,filt_pattern = ["*.xlsx"]):
    #Complete the import of the file, be sure to check if connection is running
      open_dialog = Gtk.FileChooserDialog(
          title="Please choose a file", parent=self, action=Gtk.FileChooserAction.OPEN
      )
      open_dialog.add_buttons(
          Gtk.STOCK_CANCEL,
          Gtk.ResponseType.CANCEL,
          Gtk.STOCK_OPEN,
          Gtk.ResponseType.OK,
      )
      filter = Gtk.FileFilter()
      filter.set_name("*.xlsx")
      for pat in filt_pattern:
          filter.add_pattern(pat)
      #open_dialog.add_filter(filter)

      response = open_dialog.run()
      if response == Gtk.ResponseType.OK:
          passthrough = param
          file_name = open_dialog.get_filename()
          self.import_pens(file_name,passthrough)
      elif response == Gtk.ResponseType.CANCEL:
        pass
        # don't bother building the window
        #self.destroy()
      open_dialog.destroy()

  def fileChooser_save(self,param,filt_pattern = ["*.xlsx"]):
    save_dialog = Gtk.FileChooserDialog("Save As", self,
                                        Gtk.FileChooserAction.SAVE,
                                        (Gtk.STOCK_OK, Gtk.ResponseType.OK,
                                          Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL))
    save_dialog.set_current_name('Pens_Export')
    filter = Gtk.FileFilter()
    filter.set_name("*.xlsx")
    for pat in filt_pattern:
        filter.add_pattern(pat)
    save_dialog.add_filter(filter)
    response = save_dialog.run()
    if response == Gtk.ResponseType.OK:
        file_name = save_dialog.get_filename()
        self.export_pens(file_name,param)
    else:
      pass
      # don't bother building the window
      #self.destroy()
    save_dialog.destroy()

  def export_pens(self,directory_name,t_filter,*args):
    file_name_suffix = 'xlsx'
    dest_filename = os.path.join(directory_name +"." + file_name_suffix)
    wb = Workbook()
    if self.pens_available: #checks if dictionary of pens is not empty
      sheet = wb.active
      sheet.title = "pens"
      for key, pen in self.pens_available.items():
        columns = list(pen.keys()) #get pen column headers
      sheet.append(columns)
      num_pens = len(self.pens_available)
      for key, pen in self.pens_available.items():
        row = list(pen.values())
        sheet.append(row)
    else:
      self.display_msg(msg="No Pens To Export")
    try:
      wb.save(filename = dest_filename)
      self.display_msg(msg="Exported {} Pens To File".format(num_pens))
    except:
      self.display_msg(msg="Save Failed")
      print('Save Failed')

  def import_pens(self,directory_name,conx_selected,*args):
    #Need to deal with only importing pens to connections which are stopped
    dest_filename = os.path.join(directory_name)
    wb = load_workbook(dest_filename)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
      header_row = row
    bad_pen = []
    r = 0
    pens = {}
    for row in ws.values:
      if r != 0:  #skip header row
        col = 0
        if len(header_row) == len(row) and (not None in row): #check for missing element in row
          pens[row[0]] = {} #Make key the pen ID value
          for i in header_row :
            try:
              ########################################NEED TO DO SOME MORE CHECKING OF BAD IMPORTING  #########
              if i == 'id' or i == 'chart_id' or i == 'weight':
                pens[row[0]][i] = int(row[col])
              elif i == 'scale_lock' or i == 'scale_auto':
                pens[row[0]][i] = bool(row[col])
              else:
                pens[row[0]][i] = str(row[col])
            except:
              bad_pen.append(row[0])
            col +=1
        else:
          bad_pen.append(row[0])
      r += 1
    if bad_pen:
      self.display_msg(msg="{} Pen(s) Were Incorectly Formated And Were Not Imported".format(str(len(bad_pen))))
    
    existing_keys = list(self.pens_available.keys())
    duplicate_pen = []
    new_pens = {}
    if pens:  #Are there any pens to import?
      for k in pens.keys():
        if k in existing_keys:
          duplicate_pen.append(pens[k]['id'])
        else:
          valid = self.check_pen_import(pens[k])
          if valid:
            new_pens[k] = pens[k]
    if new_pens:
      for k in new_pens.keys():
        new = self.Tbl(chart_id = new_pens[k]['chart_id'],id = new_pens[k]['id'])
        self.db_session.add(new)
        self.db_session.commit()    
        self.db_session.refresh(new)  #Retrieves newly created pen id from the database (new.id)
        row = Pen_row(new_pens[k],self.pen_grid,self.pen_row_num,self.app,self) #add row to popup
        self.create_delete_button(new_pens[k]['id'],self.pen_row_num)           #add buttons to row
        self.pen_row_num += 1
        self.pen_settings.append(row)
        self.create_pen_object(new_pens[k]['id'],new_pens[k]['chart_id'])     #add pen object to specific chart
        row.update_db()                                                       #update pen with imported settings
    self.show_all()
    self.display_msg(msg="{} Pen(s) Were Were Imported".format(str(len(new_pens.keys()))))

  def check_pen_import(self,pen,*args):
    #{'id': 10, 'chart_id': 1, 'tag_id': 'Speed', 'connection_id': 'Turbine', 'visible': 1, 'weight': '2.0', 
    # 'color': '#865e3c', 'scale_minimum': '0.0', 'scale_maximum': '50.0', 'scale_lock': 1, 'scale_auto': 1}
    valid = True
    try:
      if int(pen['chart_id'])>=1 and int(pen['chart_id'])<=16:
        pass
      else:
        valid = False
      if pen['connection_id'] in self.connections_list:
        pass
      else:
        valid = False
      if pen['visible'].isdigit() and pen['weight'].isdigit() and pen['scale_minimum'].isdigit() and pen['scale_maximum'].isdigit():
        pass
      else:
        valid = False
      if type(pen['scale_lock']) == bool and type(pen['scale_auto']) == bool:
        pass
      else:
        valid = False
    except:
      valid = False
    return valid

  def display_msg(self,msg,*args):
    popup = PopupMessage(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      return True
    else:
      return False

  def close_popup(self, button,msg="Abandon Pen Settings Changes?"):
    if self.unsaved_changes_present:
      popup = PopupConfirm(self, msg=msg)
      response = popup.run()
      popup.destroy()
      if response == Gtk.ResponseType.YES:
        self.destroy()
        return True
      else:
        return False
    else:
      self.destroy()


class Pen_row(object):

  def __init__(self,params,pen_grid,row_num,app,parent,*args):
    self.app = app    
    self.parent = parent
    self.db_settings_session = self.app.settings_db.session
    self.db_settings_model = self.app.settings_db.models['pen']
    self.Pen_Settings_Tbl = self.db_settings_model
    self.params = params
    self.pen_grid = pen_grid
    self.pen_row_num = row_num
    self.id = self.params['id']
    self.chart_id = self.params['chart_id']
    if self.params['connection_id'] == None:
      self.db_conn_id = 0
    else:
      self.db_conn_id = self.params['connection_id']
    if self.params['tag_id'] == None:
      self.db_tag_id = 0
    else:
      self.db_tag_id = self.params['tag_id']
    self.unsaved_changes = False      #Need to pass this up so that confirm closing popup with unsaved changes
    self.connections_available = {}
    self.tags_available = {}
    self.get_available_connections()
    self.get_available_tags(self.db_conn_id)
    self.build_row()
    #Rows start at 1 because row 0 is titles
    #Grid : Left,Top,Width,Height

  def build_row(self,*args):
    #Chart Select
    db_chart_number = str(self.params['chart_id'])
    selections = []
    #for num in range(self.app.charts_number):
    for num in range(16):
      selections.append(str(num+1))
    self.chart_number = Gtk.ComboBoxText(width_request = 20)
    for x in selections:
        self.chart_number.append_text(x)
    if int(db_chart_number) > 16 or int(db_chart_number) <1:
      idx = 0
    else:
      idx = selections.index(str(db_chart_number))
    sc = self.chart_number.get_style_context()
    sc.add_class('ctrl-combo')
    self.chart_number.set_active(idx)
    self.pen_grid.attach(self.chart_number,0,self.pen_row_num,1,1)
    self.chart_number.connect("changed", self.row_changed)

    #Display Status
    db_display_status = bool(self.params['visible'])
    box= Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale(os.path.join(PUBLIC_DIR, 'images/Check.png'), 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    wid =CheckBoxWidget(30,30,image,db_display_status)
    self.display_status = wid.return_self()
    sc = self.display_status.get_style_context()
    sc.add_class('check-box')
    box.set_center_widget(self.display_status)
    self.pen_grid.attach(box,1,self.pen_row_num,1,1)
    self.display_status.connect("toggled", self.row_changed)

    #Connection Select
    db_conx = str(self.params['connection_id'])
    self.conn_select = Gtk.ComboBoxText(width_request = 200, halign = Gtk.Align.CENTER )
    i = 0
    for key, val in self.connections_available.items():
      self.conn_select.append_text(val['id'])
      if val['id'] == db_conx:
        i = key
    self.conn_select.set_active(i)
    sc = self.conn_select.get_style_context()
    sc.add_class('ctrl-combo')
    self.pen_grid.attach(self.conn_select,2,self.pen_row_num,1,1)
    self.conn_select.connect("changed", self.row_changed)
    self.conn_select.connect("changed",self.new_connection_selelcted)

    #Tag Select
    db_tag = str(self.params['tag_id'])
    i = 0
    self.tag_select = Gtk.ComboBoxText(hexpand = True)
    self.tag_select.append_text("")   #Add blank selection
    if self.params['connection_id'] in self.tags_available.keys():
      for key, val in self.tags_available[self.params['connection_id']].items():
        self.tag_select.append_text(val['id'])
        if val['id'] == db_tag:
          i = key
      self.tag_select.set_active(i)
    sc = self.tag_select.get_style_context()
    sc.add_class('ctrl-combo')
    self.pen_grid.attach(self.tag_select,3,self.pen_row_num,1,1)
    self.tag_select.connect("changed", self.row_changed)

    #color
    db_color = str(self.params['color']) #example:#0000FF
    rgbcolor = Gdk.RGBA()
    rgbcolor.parse(db_color)
    rgbcolor.to_string()
    self.color_button = Gtk.ColorButton(width_request = 20)
    self.color_button.set_rgba (rgbcolor)
    sc = self.color_button.get_style_context()
    sc.add_class('ctrl-button')
    self.pen_grid.attach(self.color_button,4,self.pen_row_num,1,1)
    self.color_button.connect('color-set',self.row_changed)

    #line width
    db_line_width = str(self.params['weight']) 
    but = Gtk.Button(width_request = 20)
    self.line_width = Gtk.Label()
    self.line_width.set_label(db_line_width)
    self.add_style(self.line_width,['borderless-num-display','font-14','text-black-color'])
    but.add(self.line_width)
    sc = but.get_style_context()
    sc.add_class('ctrl-button')
    but.connect('clicked',self.open_numpad,self.line_width,{'min':0,'max':16,'type':int,'polarity':False,'name':'Line Width'})
    self.pen_grid.attach(but,5,self.pen_row_num,1,1)
    but.connect('clicked',self.row_changed)

    #scale minimum
    db_scale_minimum = str(self.params['scale_minimum']) 
    but = Gtk.Button(width_request = 100)
    self.scale_minimum = Gtk.Label()
    self.scale_minimum.set_label(db_scale_minimum)
    self.add_style(self.scale_minimum,['borderless-num-display','font-14','text-black-color'])
    but.add(self.scale_minimum)
    sc = but.get_style_context()
    sc.add_class('ctrl-button')
    but.connect('clicked',self.open_numpad,self.scale_minimum,{'min':-32768,'max':32768,'type':float,'polarity':True,'name':'Scale Minimum'})
    self.pen_grid.attach(but,6,self.pen_row_num,1,1)
    but.connect('clicked',self.row_changed)


    #scale maximum
    db_scale_maximum = str(self.params['scale_maximum']) 
    but = Gtk.Button(width_request = 100)
    self.scale_maximum = Gtk.Label()
    self.scale_maximum.set_label(db_scale_maximum)
    self.add_style(self.scale_maximum,['borderless-num-display','font-14','text-black-color'])
    but.add(self.scale_maximum)
    sc = but.get_style_context()
    sc.add_class('ctrl-button')
    but.connect('clicked',self.open_numpad,self.scale_maximum,{'min':-32768,'max':32768,'type':float,'polarity':True,'name':'Scale Maximum'})
    self.pen_grid.attach(but,7,self.pen_row_num,1,1)
    but.connect('clicked',self.row_changed)

    #Autoscale Status
    db_autoscale_status = bool(self.params['scale_auto'])
    box= Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale(os.path.join(PUBLIC_DIR, 'images/Check.png'), 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    wid =CheckBoxWidget(30,30,image,db_autoscale_status)
    self.autoscale_status = wid.return_self()
    sc = self.autoscale_status.get_style_context()
    sc.add_class('check-box')
    box.set_center_widget(self.autoscale_status)
    self.pen_grid.attach(box,8,self.pen_row_num,1,1)
    self.autoscale_status.connect("toggled", self.row_changed)

    #Lockscale Status
    db_lockscale_status = bool(self.params['scale_lock'])
    box= Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale(os.path.join(PUBLIC_DIR, 'images/Check.png'), 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    wid =CheckBoxWidget(30,30,image,db_lockscale_status)
    self.lockscale_status = wid.return_self()
    sc = self.lockscale_status.get_style_context()
    sc.add_class('check-box')
    box.set_center_widget(self.lockscale_status)
    self.pen_grid.attach(box,9,self.pen_row_num,1,1)
    self.lockscale_status.connect("toggled", self.row_changed)

    #Save Button
    self.save_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Save.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.save_button.add(image)
    sc = self.save_button.get_style_context()
    sc.add_class('ctrl-button')
    self.pen_grid.attach(self.save_button,10,self.pen_row_num,1,1)
    self.save_button.connect('clicked',self.save_settings)
    
  def open_numpad(self,button,widget_obj,params,*args):
    numpad = ValueEnter(self,widget_obj,params)
    response = numpad.run()
    if response == Gtk.ResponseType.NO:
      pass
    else:
      pass
    numpad.destroy()
  
  def row_changed(self,*args):
    self.add_style(self.save_button,['exit-button'])
    self.parent.unsaved_changes(True,self,self.id)
  
  def row_updated(self,*args):
    self.add_style(self.save_button,['ctrl-button'])
    self.parent.unsaved_changes(False,self,self.id)

  def new_connection_selelcted(self, *args):
    c_temp = self.conn_select.get_active_text()
    self.tag_select.remove_all()
    self.tag_select.append_text("")
    if c_temp in self.tags_available.keys():
      for key, val in self.tags_available[c_temp].items():
        self.tag_select.append_text(val['id'])
    self.tag_select.set_active(0)

  def save_settings(self,button,*args):
    self.update_db()
    self.row_updated()

  def update_db(self,*args):
    p_settings = {}
    p_settings['id'] = self.chart_id
    settings = self.db_settings_session.query(self.Pen_Settings_Tbl).filter(self.Pen_Settings_Tbl.id == self.id).first()  # save the current settings
    if settings:
      chart_id= int(self.chart_number.get_active_text())
      settings.chart_id = chart_id
      p_settings['chart_id'] = chart_id

      #get tag ID
      t_id = self.tag_select.get_active_text()
      settings.tag_id = t_id
      p_settings['tag_id'] = t_id

      #get connection ID
      c_id = self.conn_select.get_active_text()
      settings.connection_id = c_id
      p_settings['connection_id'] = c_id

      visible = int(self.display_status.get_active())
      settings.visible = visible
      p_settings['visible'] = visible

      weight = self.line_width.get_label()
      settings.weight = weight
      p_settings['weight'] = weight

      rgba_color = self.color_button.get_rgba()
      red = int(rgba_color.red*255)
      green = int(rgba_color.green*255)
      blue = int(rgba_color.blue*255)
      hex_color =  '#{r:02x}{g:02x}{b:02x}'.format(r=red,g=green,b=blue)
      settings.color = hex_color
      p_settings['color'] = hex_color
      
      minimum = self.scale_minimum.get_label()
      settings.scale_minimum = minimum
      p_settings['scale_minimum'] = minimum

      maximum = self.scale_maximum.get_label()
      settings.scale_maximum = maximum
      p_settings['scale_maximum'] = maximum

      lock = int(self.lockscale_status.get_active())
      settings.scale_lock = lock
      p_settings['scale_lock'] = lock

      auto = int(self.autoscale_status.get_active())
      settings.scale_auto = auto
      p_settings['scale_auto'] = auto

    self.db_settings_session.commit()
    self.update_pen_object(p_settings)

  def update_pen_object(self,p_settings,*args):
    self.app.charts[self.chart_id].pens[self.id]._chart_id = p_settings['chart_id']
    self.app.charts[self.chart_id].pens[self.id]._tag_id = p_settings['tag_id']
    self.app.charts[self.chart_id].pens[self.id]._connection_id = p_settings['connection_id']
    self.app.charts[self.chart_id].pens[self.id]._visible = p_settings['visible']
    self.app.charts[self.chart_id].pens[self.id]._weight = p_settings['weight']
    self.app.charts[self.chart_id].pens[self.id]._color = p_settings['color']
    self.app.charts[self.chart_id].pens[self.id]._scale_minimum = p_settings['scale_minimum']
    self.app.charts[self.chart_id].pens[self.id]._scale_maximum = p_settings['scale_maximum']
    self.app.charts[self.chart_id].pens[self.id]._scale_lock = p_settings['scale_lock']
    self.app.charts[self.chart_id].pens[self.id]._scale_auto = p_settings['scale_auto']
    p_obj = self.app.charts[self.chart_id].pens[self.id]

    if p_settings['chart_id'] != self.chart_id:
      #chart ID was changed so need to move pen object into other chart object
      self.app.charts[p_settings['chart_id']].pens[self.id] = p_obj
      del self.app.charts[self.chart_id].pens[self.id]
  
  def add_style(self, item,style):
    sc = item.get_style_context()
    for items in sc.list_classes():
      #remove all default styles
      sc.remove_class(items)
    for sty in style:
      #add new styles
      sc.add_class(sty)

  def get_available_connections(self,*args):

    conx_items = ['id', 'connection_type', 'description']
    new_params = {}
    count = 1
    self.connections_available = {0: {'id': '', 'connection_type': 0, 'description': ''}}
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      for c in conx_items:
        new_params[c] = getattr(conx_obj, c)
      self.connections_available[count] = new_params
      new_params = {}
      count += 1

  def get_available_tags(self,c_id,*args):
    self.tags_available = {}
    new_params = {}
    count = 1
    self.conx_tags = {}
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      if conx_obj:
        tag_items = conx_obj.return_tag_parameters()  #return list of tag parameters from the specific connection
        for tag_id,tag_obj in conx_obj.get('tags').items():
          for c in tag_items:
            new_params[c] = getattr(tag_obj, c)
          self.conx_tags[count] = new_params
          new_params = {}
          count += 1
        self.tags_available[conx_id]= self.conx_tags
        self.conx_tags = {}
        count = 1


class TagMainPopup(Gtk.Dialog):

  def __init__(self, parent,app):
    super().__init__(transient_for = parent,flags=0) 
    self.unsaved_changes_present = False
    self.unsaved_conn_rows = {}
    self.tags_available = {}
    self.connections_available = {}
    self.app = app
    self.build_window()
    self.content_area = self.get_content_area()
    self.get_available_tags('c_id')
    self.get_available_connections()
    self.tag_filter_val = ''

    self.dialog_window = Gtk.Box(width_request=800,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window)
    ### -title bar- ####
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=1050)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    ### -content area- ####
    self.base_area = Gtk.Box(spacing = 10,orientation=Gtk.Orientation.VERTICAL,margin = 20)
    self.scroll = Gtk.ScrolledWindow(width_request = 850,height_request = 600)
    self.scroll.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
    self.scroll.add(self.base_area)
    self.dialog_window.pack_start(self.scroll,1,1,1)
    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=800)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)

    self.build_header("Tag Browser")
    self.build_base()
    self.build_footer()
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(1050, 800)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(False)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self,title):
    #header
    self.tag_sort = Gtk.ComboBoxText()
    self.tag_sort.set_entry_text_column(0)
    for x in self.connections_available:
      self.tag_sort.append_text(self.connections_available[x]['id'])
    self.tag_sort.set_active(0)
    sc = self.tag_sort.get_style_context()
    sc.add_class('ctrl-combo')
    self.tag_sort.connect("changed", self.filter_tags)
    self.title_bar.pack_start(self.tag_sort,0,0,1)
    self.add_button2 = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/AddTag.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.add_button2.add(image)
    sc = self.add_button2.get_style_context()
    sc.add_class('ctrl-button')
    self.title_bar.pack_start(self.add_button2,0,0,0)
    self.add_button2.connect('clicked',self.add_tag_popup,None,self.connections_available)

    tag_import = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Import.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    tag_import.add(image)
    sc = tag_import.get_style_context()
    sc.add_class('ctrl-button')
    self.title_bar.pack_start(tag_import,0,0,0)
    tag_import.connect('clicked',self.import_popup)

    tag_export = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Export.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    tag_export.add(image)
    sc = tag_export.get_style_context()
    sc.add_class('ctrl-button')
    self.title_bar.pack_start(tag_export,0,0,0)
    tag_export.connect('clicked',self.export_popup)

    title = Gtk.Label(label=title,width_request = 500)
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-18')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    self.pin_button = Gtk.Button(width_request = 20)
    self.pin_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.pin_button.add(image)
    self.title_bar.pack_end(self.pin_button,0,0,1)
    sc = self.pin_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self):
    self.connection_settings = []  
    self.liststore = Gtk.ListStore(GdkPixbuf.Pixbuf, str , str, str , str,GdkPixbuf.Pixbuf,GdkPixbuf.Pixbuf)
    self.treeview = Gtk.TreeView(self.liststore)
    self.treeview.connect('button-press-event' , self.tree_item_clicked)
    self.treeview.set_rules_hint( True )
    self.add_style(self.treeview,['treeview'])

    #Generate Columns
    columns = {0:{'name':'','cell':Gtk.CellRendererPixbuf(),'width':30,'expand':False,'type':'pixbuf'},
               1:{'name':'Tagname','cell':Gtk.CellRendererText(),'width':-1,'expand':True,'type':'text'},
               2:{'name':'Connection','cell':Gtk.CellRendererText(),'width':-1,'expand':True,'type':'text'},
               3:{'name':'Address','cell':Gtk.CellRendererText(),'width':-1,'expand':True,'type':'text'},
               4:{'name':'Description','cell':Gtk.CellRendererText(),'width':-1,'expand':True,'type':'text'},
              }
    for c in columns:
      col = Gtk.TreeViewColumn(columns[c]['name'])
      self.treeview.append_column(col)
      col.pack_start(columns[c]['cell'], columns[c]['expand'])
      # Allow sorting on the column
      col.set_sort_column_id(c)
      if columns[c]['type'] == 'pixbuf':
        col.set_attributes(columns[c]['cell'],pixbuf=c)
        col.set_max_width(columns[c]['width'])
      else:
        col.set_attributes(columns[c]['cell'],text=c)
        col.set_expand(True)

    #Add settings button setup
    self.cell_settings = Gtk.CellRendererPixbuf()                         # create a CellRenderers to render the data
    self.tvcolumn_settings = Gtk.TreeViewColumn('')
    self.treeview.append_column(self.tvcolumn_settings)
    self.tvcolumn_settings.pack_end(self.cell_settings, False)
    self.tvcolumn_settings.set_attributes(self.cell_settings,pixbuf=5)
    self.tvcolumn_settings.set_max_width(30)
    #Add delete button setup
    self.cell_delete = Gtk.CellRendererPixbuf()
    self.tvcolumn_delete = Gtk.TreeViewColumn('')
    self.treeview.append_column(self.tvcolumn_delete)
    self.tvcolumn_delete.pack_end(self.cell_delete, False)
    self.tvcolumn_delete.set_attributes(self.cell_delete,pixbuf=6)
    self.tvcolumn_delete.set_max_width(30)

    # make treeview searchable
    self.treeview.set_search_column(2)
    # Allow drag and drop reordering of rows
    self.treeview.set_reorderable(True)
    #Add treeview to base window
    self.base_area.add(self.treeview)

    #header
    self.add_tag_rows(self.tag_filter_val)
    self.show_all()

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.save_settings,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    #self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    self.apply_button.connect('clicked',self.save_settings,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    #self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def import_tags(self,directory_name,conx_selected,*args):
    #Need to deal with only importing tags to connections which are stopped
    dest_filename = os.path.join(directory_name)
    wb = load_workbook(dest_filename)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
      header_row = row
    bad_tag = []
    r = 0
    tags = {}
    for row in ws.values:
      if r != 0:  #skip header row
        tags[r] = {}
        col = 0
        if row[1] == conx_selected:           #check if tags connection_id is the same is the one selected
          if len(header_row) == len(row) and (not None in row): #check for missing element in row
            for i in header_row :
              tags[r][i] = str(row[col])
              col +=1
          else:
            bad_tag.append(row[0])
      r += 1
    if bad_tag:
      self.display_msg(msg="{} Tag(s) Were Incorectly Formated And Were Not Imported".format(str(len(bad_tag))))
    
    duplicate_tag = []
    new_tags = []
    existing_tags = []
    if tags:  #Are there any tags to import?
      ##$#conx_obj = self.app.link.get('connections').get(conx_selected)
      conx_obj = self.app.db.get('connections').get(conx_selected)
      if conx_obj != None:
        for tag_id,tag_obj in conx_obj.get('tags').items():
            existing_tags.append(tag_id)      #Collect all tags in
        for k in tags.keys():
          if tags[k]['id'] in existing_tags:
            duplicate_tag.append(tags[k]['id'])
          else:
            self.create_tag(tags[k])
            new_tags.append(tags[k]['id'])
     
    if duplicate_tag and not new_tags:
      self.display_msg(msg="{} Tag(s) Already Exist And Were Not Imported".format(str(len(duplicate_tag))))
    elif new_tags:
      self.display_msg(msg="{} Tag(s) Were Imported".format(str(len(new_tags))))
    else:
      self.display_msg(msg="No Tags Imported")

  def export_tags(self,directory_name,t_filter,*args):
    file_name_suffix = 'xlsx'
    dest_filename = os.path.join(directory_name +"." + file_name_suffix)
    wb = Workbook()
    if self.tags_available[t_filter]: #checks if dictionary of tags is not empty
      sheet = wb.active
      sheet.title = t_filter
      for num in self.tags_available[t_filter]:
          columns = list(self.tags_available[t_filter][num].keys()) #get tag column headers
      sheet.append(columns)
      num_tags = len(self.tags_available[t_filter])
      for x in range(len(self.tags_available[t_filter])): #fill table with number of tags available in connection
        row = list(self.tags_available[t_filter][x+1].values())
        #sheet.cell(column=c, row=(x+2), value=self.tags_available[t_filter][x+1][header])
        sheet.append(row)
    else:
      self.display_msg(msg="No Tags In Connection To Export")
    try:
      wb.save(filename = dest_filename)
      self.display_msg(msg="Exported {} Tags To File".format(num_tags))
    except:
      self.display_msg(msg="Save Failed")
      print('Save Failed')

  def tree_item_clicked(self, treeview, event):
    pthinfo = treeview.get_path_at_pos(event.x, event.y)
    if event.button == 1: #left click
      if pthinfo != None:
        path,column,cellx,celly = pthinfo
        treeview.grab_focus()
        treeview.set_cursor(path,column,0)
        print(column.get_title())
        #update currently active display
        selection = treeview.get_selection()
        tree_model, tree_iter = selection.get_selected()
        #If selected column is delete icon then initiate delete of tag
        if tree_iter != None:
          #gathers the Tag name/Connection column text in the row clicked on
          t_id = tree_model[tree_iter][1]
          c_id = tree_model[tree_iter][2]
          #checks if it is a delete or settings button click
          if column is self.tvcolumn_delete:
            self.confirm_delete('',t_id,c_id,tree_iter)
          elif column is self.tvcolumn_settings:
            self.open_settings_popup(t_id,c_id)
      else:
        #unselect row in treeview
        selection = treeview.get_selection()
        selection.unselect_all()
    elif event.button == 3: #right click
      if pthinfo != None:
        path,col,cellx,celly = pthinfo
        treeview.grab_focus()
        treeview.set_cursor(path,col,0)
        rect = Gdk.Rectangle()
        rect.x = event.x
        rect.y = event.y + 10
        rect.width = rect.height = 1
        selection = treeview.get_selection()
        tree_model, tree_iter = selection.get_selected()
        popover = Gtk.Popover(width_request = 200)
        vbox = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
        if tree_iter is not None:
          #gathers the Tag name/Connection column text in the row clicked on
          t_id = tree_model[tree_iter][1]
          c_id = tree_model[tree_iter][2]
          #popover to add display
          edit_btn = Gtk.ModelButton(label="Edit", name=t_id)
          #cb = lambda btn: self.open_widget_popup(btn)
          #edit_btn.connect("clicked", cb)
          vbox.pack_start(edit_btn, False, True, 10)
          delete_btn = Gtk.ModelButton(label="Delete", name=t_id)
          cb = lambda btn:self.confirm_delete('',t_id,c_id,tree_iter)
          delete_btn.connect("clicked", cb)
          vbox.pack_start(delete_btn, False, True, 10)
        popover.add(vbox)
        popover.set_position(Gtk.PositionType.RIGHT)
        popover.set_relative_to(treeview)
        popover.set_pointing_to(rect)
        popover.show_all()
        sc = popover.get_style_context()
        sc.add_class('popover-bg')
        sc.add_class('font-16')
        return
      else:
        return
    selection = treeview.get_selection()
    tree_model, tree_iter = selection.get_selected()

  def add_tag_rows(self,filter,*args):
    tag_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Tag.png', 25, 25)
    settings_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/settings.png', 25, 25)
    delete_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Delete.png', 25, 25)
    for tags in self.tags_available:
      for tag in self.tags_available[tags]:
        conx_id = tags
        if filter == '' or filter == conx_id:
          #Need a better way to deal with this
          if 'address' in self.tags_available[tags][tag].keys():
            self.liststore.append([tag_icon,
                                  self.tags_available[tags][tag]['id'],
                                  self.tags_available[tags][tag]['connection_id'],
                                  str(self.tags_available[tags][tag]['address']),
                                  self.tags_available[tags][tag]['description'],
                                  settings_icon,
                                  delete_icon
                                            ])
          else:
            self.liststore.append([tag_icon,
                                   self.tags_available[tags][tag]['id'],
                                   self.tags_available[tags][tag]['connection_id'],
                                   '',
                                   self.tags_available[tags][tag]['description'],
                                   settings_icon,
                                   delete_icon
                                            ])
    self.show_all()

  def add_column_names(self,*args):
    labels = ['','Tag Name', 'Connection', 'Description','Address','',''] # may want to create a table in the db for column names
    for l_idx in range(len(labels)):
        l = Gtk.Label(labels[l_idx])
        sc = l.get_style_context()
        sc.add_class('text-black-color')
        sc.add_class('font-14')
        sc.add_class('font-bold')
        #self.grid.attach(l, l_idx, 0, 1, 1)
    l_row = Gtk.ListBoxRow()
    hbox = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, height_request = 30)
    l_row.add(hbox)
    for l in labels:
      if l == '':
        lbl = Gtk.Label(label=l, xalign=0.5, yalign = 0.5, width_request = 20)
      else:     
        lbl = Gtk.Label(label=l, xalign=0.5, yalign = 0.5, width_request = 175)         
      hbox.pack_start(lbl, True, True, 0)
      self.add_style(lbl,['font-16','font-bold','text-black-color'])
    self.listbox.add(l_row)

  def get_available_tags(self,c_id,*args):
    new_params = {}
    count = 1
    self.conx_tags = {}
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      tag_items = conx_obj.return_tag_parameters()  #return list of tag parameters from the specific connection
      for tag_id,tag_obj in conx_obj.get('tags').items():
        for c in tag_items:
          new_params[c] = getattr(tag_obj, c)
        self.conx_tags[count] = new_params
        new_params = {}
        count += 1
      self.tags_available[conx_id]= self.conx_tags
      self.conx_tags = {}
      count = 1

  def get_available_connections(self,*args):

    conx_items = ['id', 'connection_type', 'description']
    new_params = {}
    count = 1
    self.connections_available = {0: {'id': '', 'connection_type': 0, 'description': ''}}
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      for c in conx_items:
        new_params[c] = getattr(conx_obj, c)
      self.connections_available[count] = new_params
      new_params = {}
      count += 1

  def filter_tags(self,*args):
    self.tag_filter_val = self.tag_sort.get_active_text()
    self.remove_all_rows()
    self.add_tag_rows(self.tag_filter_val)

  def remove_all_rows(self,*args):
    self.liststore.clear()

  def delete_row(self,t_id,c_id,*args):
    ##$#conx_obj = self.app.link.get("connections").get(c_id)
    conx_obj = self.app.db.get("connections").get(c_id)
    if conx_obj != None:
      tag_obj = conx_obj.get('tags').get(t_id)
      self.app.link.delete_tag(tag_obj,t_id,c_id)
    self.show_all()

  def confirm_delete(self, button,tag_id,conx_id,tree_iter,msg="Are you sure you want to delete this tag?", args=[]):
    popup = PopupConfirm(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      self.delete_row(tag_id,conx_id)
      self.liststore.remove(tree_iter)
      return True
    else:
      return False

  def export_popup(self,*args):
    popup = Export_ImportTagsPopup(self, self.app,export = True)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      results = (popup.get_result())
      self.fileChooser_save(results['conx_select'],filt_pattern = ["*.xlsx"])
    else:
      return False

  def import_popup(self,*args):
    popup = Export_ImportTagsPopup(self, self.app,export = False)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      results = (popup.get_result())
      self.fileChooser_open(results['conx_select'],filt_pattern = ["*.xlsx"])
    else:
      return False

  def fileChooser_open(self,conx_sel,filt_pattern = ["*.xlsx"]):
    #Complete the import of the file, be sure to check if connection is running
      open_dialog = Gtk.FileChooserDialog(
          title="Please choose a file", parent=self, action=Gtk.FileChooserAction.OPEN
      )
      open_dialog.add_buttons(
          Gtk.STOCK_CANCEL,
          Gtk.ResponseType.CANCEL,
          Gtk.STOCK_OPEN,
          Gtk.ResponseType.OK,
      )
      filter = Gtk.FileFilter()
      filter.set_name("*.xlsx")
      for pat in filt_pattern:
          filter.add_pattern(pat)
      open_dialog.add_filter(filter)

      response = open_dialog.run()
      if response == Gtk.ResponseType.OK:
          conx_name_selected = conx_sel
          file_name = open_dialog.get_filename()
          self.import_tags(file_name,conx_name_selected)
      elif response == Gtk.ResponseType.CANCEL:
          # don't bother building the window
          self.destroy()
      open_dialog.destroy()

  def fileChooser_save(self,conx_sel,filt_pattern = ["*.xlsx"]):
    save_dialog = Gtk.FileChooserDialog("Save As", self,
                                        Gtk.FileChooserAction.SAVE,
                                        (Gtk.STOCK_OK, Gtk.ResponseType.OK,
                                          Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL))
    save_dialog.set_current_name('Tags_Export')
    filter = Gtk.FileFilter()
    filter.set_name("*.xlsx")
    for pat in filt_pattern:
        filter.add_pattern(pat)
    save_dialog.add_filter(filter)
    response = save_dialog.run()
    if response == Gtk.ResponseType.OK:
        tag_filter = conx_sel
        file_name = save_dialog.get_filename()
        self.export_tags(file_name,tag_filter)
    else:
        # don't bother building the window
        self.destroy()
    save_dialog.destroy()

  def add_tag_popup(self,button,duplicate_name_params,*args):
    popup = AddTagPopup(self,duplicate_name_params,self.app,self.connections_available)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      #results = (popup.get_result())
      #self.check_duplicate_name(results)
      return True
    else:
      return False

  def get_tag_params(self,tag_id,conx_id):
    new_params = {}
    ##$#conx_obj = self.app.link.get("connections").get(conx_id)
    conx_obj = self.app.db.get("connections").get(conx_id)
    if conx_obj != None:
      tag_items = conx_obj.return_tag_parameters()  #return list of tag parameters from the specific connection
      tag_obj = conx_obj.get('tags').get(tag_id)
      if tag_obj != None:
        for c in tag_items:
          new_params[c] = getattr(tag_obj, c)
        return new_params

  def check_duplicate_name(self,results,*args):
    dup = False
    ##$#conx_obj = self.app.link.get('connections').get(results['connection_id'])
    conx_obj = self.app.db.get('connections').get(results['connection_id'])
    if conx_obj != None:
      for tag_id,tag_obj in conx_obj.get('tags').items():
          if tag_id == results['id']:
            dup = True
    if dup:
      self.add_tag_popup(None,results,self.connections_available)
    else:
      self.create_tag(results)
      self.open_settings_popup(results['id'],results['connection_id'])

  def save_settings(self,button,auto_close,*args):
    pass

  def open_settings_popup(self,tag_id,conx_id,*args):
    params = self.get_tag_params(tag_id,conx_id)
    popup = TagSettingsPopup(self,params,self.app)
    response = popup.run()
    popup.destroy()
    self.remove_all_rows()
    self.get_available_tags('c_id')
    self.add_tag_rows(self.tag_filter_val)
    self.show_all()

  def display_msg(self,msg,*args):
    popup = PopupMessage(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      return True
    else:
      return False

  def create_tag(self,params,*args):
    if 'address' not in params:
      params['address'] = '12'  # default address when not passed in because required to create tag
    ##$#conx_obj = self.app.link.get('connections').get(params['connection_id'])
    conx_obj = self.app.db.get('connections').get(params['connection_id'])
    conx_obj.new_tag({"id": params['id'],
                            "connection_id": params['connection_id'],
                            "description": params['description'],
                            "datatype": params['datatype'],
                            "address": params['address'],
    })
    tag_obj = conx_obj.get('tags').get(params['id'])
    if tag_obj != None:
      self.app.link.save_tag(tag_obj)
    params = self.get_tag_params(params['id'],params['connection_id'])
    self.insert_tag_row(None,params)

  def update_tag(self,params,*args):
    ##$#conx_obj = self.app.link.get('connections').get(params['connection_id'])
    conx_obj = self.app.db.get('connections').get(params['connection_id'])
    if conx_obj != None:
      tag_obj = conx_obj.get('tags').get(params['id'])
      if tag_obj != None:
        for key, val in params.items():
          if key == 'id' or key == 'connection_id' or val == None:
            pass
          else:
            try:
              tag_obj.set(key,val)
            except KeyError as e:
              print(e,key)
        self.app.link.save_tag(tag_obj)

  def insert_tag_row(self,button,params,*args):
    tag_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Tag.png', 25, 25)
    settings_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/settings.png', 25, 25)
    delete_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Delete.png', 25, 25)
    if 'address' in params.keys():
      address = params['address']
    else:
      address = ''
    self.liststore.insert(0, [tag_icon,params['id'], params['connection_id'],address, params['description'],settings_icon,delete_icon])
    self.show_all()

  def add_style(self, wid, style):
    #style should be a list
    sc = wid.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def close_popup(self, button):
    self.destroy()


class AddTagPopup(Gtk.Dialog):
  def __init__(self, parent,params,app,conx_type):
    Gtk.Dialog.__init__(self, '',parent, Gtk.DialogFlags.MODAL,
                        ()
                        )
    self.parent = parent
    self.app = app
    self.params = params
    self.conx_type = conx_type
    self.datatypes = ['UINT','INT','REAL','DINT','UDINT','BOOL','SINT','USINT']
    self.build_window()
    self.connect("response", self.on_response)
    self.result = {}

    self.content_area = self.get_content_area()
    self.dialog_window = Gtk.Box(width_request=600,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window )
    ### - Title Bar- ###
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)

    ### - Base Area- ###
    self.grid = Gtk.Grid(column_spacing=4, row_spacing=4, column_homogeneous=True, row_homogeneous=True,)
    self.dialog_window.pack_start(self.grid,1,1,1)

    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)
    self.load_settings()
    self.build_header()
    self.build_base()
    self.build_footer()
    if self.params:
      self.reload_popup(self.params)
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(200, 150)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(True)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self,*args):
    #Save Button
    self.save_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Save.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.save_button.add(image)
    sc = self.save_button.get_style_context()
    sc.add_class('ctrl-button')
    bx = Gtk.Box()
    bx.pack_end(self.save_button,0,0,0)
    #self.title_bar.pack_start(bx,0,0,0)
    self.save_button.connect('clicked',self.save_settings,False)

    #title
    title = Gtk.Label(label='Create New Tag')
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-18')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    #exit button
    self.exit_button = Gtk.Button(width_request = 20)
    self.exit_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.exit_button.add(image)
    #self.title_bar.pack_end(self.exit_button,0,0,1)
    sc = self.exit_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self,*args):
    self.pop_lbl = Gtk.Label('')
    self.add_style(self.pop_lbl,['text-red-color','font-14','font-bold'])
    self.grid.attach(self.pop_lbl,0,0,3,1)

    #Tag name entry
    lbl = Gtk.Label('Tag Name')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,1,1,1) 
    self.tag_name = Gtk.Entry(max_length = 100,width_request = 300,height_request = 30)
    self.tag_name.set_placeholder_text('Enter Tag Name')
    self.tag_name.set_alignment(0.5)
    self.add_style(self.tag_name,["entry","font-12"])
    self.tag_name.connect("notify::text-length", self.enable_new)
    self.grid.attach(self.tag_name,1,1,2,1)    

    #Connection Driver
    lbl = Gtk.Label('Connection Driver')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,2,1,1) 
    self.conx_driver = Gtk.ComboBoxText(width_request = 200,height_request = 30)#hexpand = True
    self.add_style(self.conx_driver,["font-18","list-select","font-bold"])
    for conx in self.conx_type:
       self.conx_driver.append_text(self.conx_type[conx]['id'])
    self.conx_driver.set_active(0)
    self.grid.attach(self.conx_driver,1,2,2,1)

   #Tag description entry
    lbl = Gtk.Label('Tag Description')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,3,1,1) 
    self.tag_descr = Gtk.Entry(max_length = 100,width_request = 300,height_request = 30)
    self.tag_descr.set_placeholder_text('Enter Tag Description')
    self.tag_descr.set_alignment(0.5)
    self.add_style(self.tag_descr,["entry","font-12"])
    self.tag_descr.connect("notify::text-length", self.enable_new)
    self.grid.attach(self.tag_descr,1,3,2,1)

    #Tag Datatype
    lbl = Gtk.Label('Tag Datatype')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,4,1,1) 
    self.tag_datatype = Gtk.ComboBoxText(width_request = 200,height_request = 30,halign = Gtk.Align.CENTER)#hexpand = True
    self.add_style(self.tag_datatype,["font-18","list-select","font-bold"])
    for dt in self.datatypes:
       self.tag_datatype.append_text(dt)
    self.tag_datatype.set_active(0)
    self.grid.attach(self.tag_datatype,1,4,2,1)    
    sep = Gtk.Label(height_request=3)
    self.dialog_window.pack_start(sep,1,1,1)

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.save_settings,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    self.apply_button.connect('clicked',self.save_settings,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    #self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def load_settings(self,*args):
    pass

  def save_settings(self,button,auto_close,*args):
    ####Get Results
    #{'id': '2', 'connection_type': 4, 'description': 'EthernetIP'}
    self.result['id'] = self.tag_name.get_text ()
    self.result['connection_id'] = self.conx_driver.get_active_text()
    self.result['description'] = self.tag_descr.get_text ()
    self.result['datatype'] = self.tag_datatype.get_active_text()
    #####Check Results /Save
    if self.result['id'] ==  '':
      dup = True
    else:
      dup = False

    ##$#conx_obj = self.app.link.get('connections').get(self.result['connection_id'])
    conx_obj = self.app.db.get('connections').get(self.result['connection_id'])
    if conx_obj != None:
      for tag_id,tag_obj in conx_obj.get('tags').items():
          if tag_id == self.result['id']:
            dup = True
    if dup:
      self.reload_popup(self.result)
      #self.add_tag_popup(None,self.result,self.connections_available)
    else:
      self.close_popup(False)
      self.parent.create_tag(self.result)
      self.parent.open_settings_popup(self.result['id'],self.result['connection_id'])

  def add_style(self, item,style):
    sc = item.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def enable_new(self, obj, prop):
    enable = (obj.get_property('text-length') > 0)
    if enable:
      self.add_style(obj,["entry","font-18","font-bold"])
    else:
      self.add_style(obj,["entry","font-12"])

  def close_popup(self, button):
    self.destroy()
  
  def on_response(self, widget, response_id):
    #{'id': '2', 'connection_type': 4, 'description': 'EthernetIP'}
    self.result['id'] = self.tag_name.get_text ()
    self.result['connection_id'] = self.conx_driver.get_active_text()
    self.result['description'] = self.tag_descr.get_text ()
    self.result['datatype'] = self.tag_datatype.get_active_text()
  
  def get_result(self):
    return self.result
  
  def reload_popup(self,results):
    self.pop_lbl.set_label('Name Already Exists')
    self.tag_name.set_text('')
    val = 0
    for conx in self.conx_type:
      self.conx_driver.append_text(self.conx_type[conx]['id'])
      if self.conx_type[conx]['id'] == results['connection_id']:
        self.conx_driver.set_active(val)
      val +=1
    self.tag_descr.set_text(results['description'])
    val = 0
    for dt in self.datatypes:
      self.tag_datatype.append_text(dt)
      if dt == results['datatype']:
        self.tag_datatype.set_active(val)
      val += 1


class TagSettingsPopup(Gtk.Dialog):
  def __init__(self, parent,params,app):
    Gtk.Dialog.__init__(self, '',parent, Gtk.DialogFlags.MODAL,
                        ()
                        )
    self.params = params
    self.app = app
    self.datatypes = ['UINT','INT','REAL','DINT','UDINT','BOOL','SINT','USINT']
    self.f_types = ['1','2','3','4']
    self.result = {}
    self.build_window()

    self.content_area = self.get_content_area()
    self.dialog_window = Gtk.Box(width_request=500,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window)
    ### -title bar- ####
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=800)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    ### - Base Area- ###
    self.grid = Gtk.Grid(column_spacing=4, row_spacing=4, column_homogeneous=True, row_homogeneous=True,)
    self.dialog_window.pack_start(self.grid,1,1,1)
    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=800)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)

    self.build_header()
    self.build_base()
    self.build_footer()
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(500, 400)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(True)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self):
    #Save Button
    self.save_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Save.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.save_button.add(image)
    sc = self.save_button.get_style_context()
    sc.add_class('ctrl-button')
    bx = Gtk.Box()
    bx.pack_end(self.save_button,0,0,0)
    #self.title_bar.pack_start(bx,0,0,0)
    self.save_button.connect('clicked',self.save_settings,False)

    #title
    tit = self.params['tag_type']
    title = Gtk.Label(label=f'{tit} - Tag')
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-18')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    #exit button
    self.exit_button = Gtk.Button(width_request = 20)
    self.exit_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.exit_button.add(image)
    #self.title_bar.pack_end(self.exit_button,0,0,1)
    sc = self.exit_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self,*args):
    row = 0
    #Tag name entry
    lbl = Gtk.Label('Tag Name')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,row,1,1) 
    self.conx_name = Gtk.Label(width_request = 300,height_request = 30)
    self.conx_name.set_text(self.params['id'])
    self.conx_name.set_alignment(0.5,0.5)
    self.add_style(self.conx_name,["label","font-18","font-bold"])
    #self.conx_name.connect("notify::text-length", self.enable_new)
    self.grid.attach(self.conx_name,1,row,2,1)   
    row+=1 

    #Tag description entry
    lbl = Gtk.Label('Tag Description')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,row,1,1) 
    #self.conx_descr = Gtk.Label(width_request = 300,height_request = 30)
    self.conx_descr = Gtk.Entry(max_length = 100,width_request = 300,height_request = 30)
    self.conx_descr.set_alignment(0.5)
    #self.add_style(self.conx_descr,["label","font-18","font-bold"])
    self.add_style(self.conx_descr,["entry","font-18","font-bold"])
    self.conx_descr.set_text(self.params['description'])
    #self.conx_descr.connect("notify::text-length", self.enable_new)
    self.grid.attach(self.conx_descr,1,row,2,1) 
    row+=1 

    #Tag Datatype
    db_dt = str(self.params['datatype'])
    lbl = Gtk.Label('Tag Datatype')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,row,1,1) 
    self.tag_datatype = Gtk.ComboBoxText(width_request = 200,height_request = 30,halign = Gtk.Align.CENTER)#hexpand = True
    self.add_style(self.tag_datatype,["font-18","list-select","font-bold"])
    found = None
    val = 0
    for dt in self.datatypes:
      self.tag_datatype.append_text(dt)
      if dt == db_dt:
         found = val
      val+= 1
    if found:
      self.tag_datatype.set_active(found)
    else:
      self.tag_datatype.set_active(0)
    self.grid.attach(self.tag_datatype,1,row,2,1)
    row+=1

    #Tag Address entry
    if 'address' in self.params.keys(): 
      db_host = str(self.params['address'])
      lbl = Gtk.Label('Tag Address')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      self.tag_address = Gtk.Entry(max_length = 100,width_request = 300,height_request = 30)
      self.tag_address.set_alignment(0.5)
      self.add_style(self.tag_address,["entry","font-18","font-bold"])
      self.tag_address.set_text(db_host)
      self.grid.attach(self.tag_address,1,row,2,1)
      row+=1 

    #Modbus function code select
    # 01-Read Only Coil , 02-R/W Coil, 03-R/W Holding Registers, 04-Read Holding Registers    
    if 'func_type' in self.params.keys():
      db_ft = str(self.params['func_type'])
      lbl = Gtk.Label('Modbus Function Code')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      self.function_type = Gtk.ComboBoxText(width_request = 200,height_request = 30,halign = Gtk.Align.CENTER)#hexpand = True
      self.add_style(self.function_type,["font-18","list-select","font-bold"])
      found = None
      val = 0
      for ft in self.f_types:
        self.function_type.append_text(ft)
        if ft == db_ft:
            found = val
        val+= 1
      if found:
        self.function_type.set_active(found)
      else:
        self.function_type.set_active(0)
      self.grid.attach(self.function_type,1,row,2,1)
      row+=1 


    #Bit
    if 'bit' in self.params.keys():
      db_tag_bit = str(self.params['bit'])
      lbl = Gtk.Label('Bit')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      self.tag_bit = Gtk.ComboBoxText(width_request = 200,height_request = 30)#hexpand = True
      self.add_style(self.tag_bit,["font-18","list-select","font-bold"])
      br = None
      val = 0
      bit_num = ['1','2','3','4','5','6','7','8','9','10','11','12','13','14','15','16']
      for item in bit_num:
        self.tag_bit.append(str(val),item)
        if item == db_tag_bit:
          br = val
        val+= 1
      if br:
        self.tag_bit.set_active(br)
      else:
        self.tag_bit.set_active(0)
      self.grid.attach(self.tag_bit,1,row,2,1)
      row+=1 

    #Word Swap
    if 'word_swapped' in self.params.keys(): 
      db_word_swapped = str(self.params['word_swapped'])
      lbl = Gtk.Label('Word Swapped')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      bx = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, halign = Gtk.Align.CENTER)
      p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale(os.path.join(PUBLIC_DIR, 'images/Check.png'), 20, -1, True)
      image = Gtk.Image(pixbuf=p_buf)
      wid =CheckBoxWidget(30,30,image,db_word_swapped)
      self.word_swapped = wid.return_self()
      bx.pack_start(self.word_swapped,0,0,0)
      self.grid.attach(bx,1,row,2,1)
      row+=1

    #Byte Swap
    if 'byte_swapped' in self.params.keys(): 
      db_byte_swapped = str(self.params['byte_swapped'])
      lbl = Gtk.Label('Byte Swapped')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      bx = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, halign = Gtk.Align.CENTER)
      p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale(os.path.join(PUBLIC_DIR, 'images/Check.png'), 20, -1, True)
      image = Gtk.Image(pixbuf=p_buf)
      wid =CheckBoxWidget(30,30,image,db_byte_swapped)
      self.byte_swapped = wid.return_self()
      bx.pack_start(self.byte_swapped,0,0,0)
      self.grid.attach(bx,1,row,2,1)
      row+=1

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.save_settings,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    self.apply_button.connect('clicked',self.save_settings,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    #self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def add_style(self, item,style):
    sc = item.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def enable_new(self, obj, prop):
    enable = (obj.get_property('text-length') > 0)
    if enable:
      self.add_style(obj,["entry","font-18","font-bold"])
    else:
      self.add_style(obj,["entry","font-12"])

  def close_popup(self, button):
    self.destroy()

  def save_settings(self,button,auto_close,*args):
    ###Get tag values
    self.result['id'] = self.conx_name.get_text()
    self.result['connection_id'] = self.params['connection_id']
    self.result['description'] = self.conx_descr.get_text ()
    if 'bit' in self.params.keys():
      self.result['bit'] = self.tag_bit.get_active_text()
    else:
      self.result['bit'] = None
    if 'address' in self.params.keys():
      self.result['address'] = str(self.tag_address.get_text())
    else:
      self.result['address'] = None
    if 'func_type' in self.params.keys():
      self.result['func_type'] = int(self.function_type.get_active_text())
    else:
      self.result['func_type'] = 4
    if 'datatype' in self.params.keys():
      self.result['datatype'] = self.tag_datatype.get_active_text()
    else:
      self.result['datatype'] = None
    if 'word_swapped' in self.params.keys():
      self.result['word_swapped'] = self.word_swapped.get_active()
    else:
      self.result['word_swapped'] = None
    if 'byte_swapped' in self.params.keys():
      self.result['byte_swapped'] = self.byte_swapped.get_active()
    else:
      self.result['byte_swapped'] = None
    ###Update tag value
    ##$#conx_obj = self.app.link.get('connections').get(self.result['connection_id'])
    conx_obj = self.app.db.get('connections').get(self.result['connection_id'])
    if conx_obj != None:
      tag_obj = conx_obj.get('tags').get(self.result['id'])
      if tag_obj != None:
        for key, val in self.result.items():
          if key == 'id' or key == 'connection_id' or val == None:
            pass
          else:
            try:
              tag_obj.set(key,val)
            except KeyError as e:
              print(e,key)
        self.app.link.save_tag(tag_obj)
    self.close_popup(False)

  def get_result(self):
    return self.result

  def open_numpad(self,button,widget_obj,params,*args):
    numpad = ValueEnter(self,widget_obj,params)
    response = numpad.run()
    if response == Gtk.ResponseType.NO:
      pass
    else:
      pass
    numpad.destroy()


class ConnectionsMainPopup(Gtk.Dialog):
  ################### get_conx_polling_status needs to be universal for all connections when calling method
  def __init__(self, parent,app):
    super().__init__(transient_for = parent,flags=0) 
    self.unsaved_changes_present = False
    self.unsaved_conn_rows = {}
    self.conn_column_names = ['id', 'connection_type', 'description']
    self.connections_available = {}
    self.tags_available = {}
    self.conx_obj_available = {}
    self.app = app
    ##$#self.conx_type = self.app.link.get('connection_types')
    self.conx_type = self.app.db.get('connection_types')
    self.build_window()
    self.content_area = self.get_content_area()
    self.conn_filter_val = ''
    self.get_available_connections()
    self.get_available_tags('')
    self.get_conx_polling_status('Turbine')

    self.dialog_window = Gtk.Box(width_request=800,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window)
    ### -title bar- ####
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=800)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    ### -content area- ####
    self.base_area = Gtk.Box(spacing = 10,orientation=Gtk.Orientation.VERTICAL,margin = 20)
    self.scroll = Gtk.ScrolledWindow(width_request = 800,height_request = 600)
    self.scroll.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
    self.scroll.add(self.base_area)
    self.dialog_window.pack_start(self.scroll,1,1,1)
    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=800)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)

    self.build_header()
    self.build_base()
    self.build_footer()
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(500, 800)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(False)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self):
    #header
    self.add_button2 = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/AddConnection.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.add_button2.add(image)
    sc = self.add_button2.get_style_context()
    sc.add_class('ctrl-button')
    self.title_bar.pack_start(self.add_button2,0,0,0)
    self.add_button2.connect('clicked',self.add_connection_popup,None,self.conx_type)

    title = Gtk.Label(label="Connection Settings",width_request = 500)
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-18')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    self.pin_button = Gtk.Button(width_request = 20)
    self.pin_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.pin_button.add(image)
    self.title_bar.pack_end(self.pin_button,0,0,1)
    sc = self.pin_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self):
    self.connection_settings = []  
    self.liststore = Gtk.ListStore(GdkPixbuf.Pixbuf,GdkPixbuf.Pixbuf,str , str, str, GdkPixbuf.Pixbuf,GdkPixbuf.Pixbuf)
    self.treeview = Gtk.TreeView(self.liststore)
    self.treeview.connect('button-press-event' , self.tree_item_clicked)
    self.treeview.set_rules_hint( True )
    self.add_style(self.treeview,['treeview'])

    #Add connection status
    self.conx_status = Gtk.CellRendererPixbuf()
    self.tvcolumn_conx_status= Gtk.TreeViewColumn('')
    self.treeview.append_column(self.tvcolumn_conx_status)
    self.tvcolumn_conx_status.pack_end(self.conx_status, False)
    self.tvcolumn_conx_status.set_attributes(self.conx_status,pixbuf=0)
    self.tvcolumn_conx_status.set_max_width(30)

    #Add connection btton
    self.conx_button = Gtk.CellRendererPixbuf()
    self.tvcolumn_conx_button= Gtk.TreeViewColumn('')
    self.treeview.append_column(self.tvcolumn_conx_button)
    self.tvcolumn_conx_button.pack_end(self.conx_button, False)
    self.tvcolumn_conx_button.set_attributes(self.conx_button,pixbuf=1)
    self.tvcolumn_conx_button.set_max_width(30)

    # #Add toggle button
    # connection_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Connect.png', 25, 25)
    # image = Gtk.Image(pixbuf=connection_icon)
    # renderer_toggle = Gtk.CellRendererToggle()
    # renderer_toggle.set_property('cell-background','gray')
    # self.tvcolumn_toggle = Gtk.TreeViewColumn('', renderer_toggle, active=1)
    # h_but = self.tvcolumn_toggle.get_button() #Get reference to column header button
    # c = h_but.get_child()
    # c.add(image)  #add image to column header button
    # c.show_all()

    #renderer_toggle.connect("toggled", self.conx_connect_toggle)
    #self.treeview.append_column(self.tvcolumn_toggle)
    #self.tvcolumn_toggle.set_max_width(30)

    #Generate Columns
    columns = {2:{'name':'Name','cell':Gtk.CellRendererText(),'width':-1,'expand':True,'type':'text'},
               3:{'name':'Connection Type','cell':Gtk.CellRendererText(),'width':-1,'expand':True,'type':'text'},
               4:{'name':'Description','cell':Gtk.CellRendererText(),'width':-1,'expand':True,'type':'text'},
              }
    for c in columns:
      col = Gtk.TreeViewColumn(columns[c]['name'])
      self.treeview.append_column(col)
      col.pack_start(columns[c]['cell'], columns[c]['expand'])
      # Allow sorting on the column
      col.set_sort_column_id(c)
      if columns[c]['type'] == 'pixbuf':
        col.set_attributes(columns[c]['cell'],pixbuf=c)
        col.set_max_width(columns[c]['width'])
      else:
        col.set_attributes(columns[c]['cell'],text=c)
        col.set_expand(True)

    #Add settings button setup
    self.cell_settings = Gtk.CellRendererPixbuf()                         # create a CellRenderers to render the data
    self.tvcolumn_settings = Gtk.TreeViewColumn('')
    self.treeview.append_column(self.tvcolumn_settings)
    self.tvcolumn_settings.pack_end(self.cell_settings, False)
    self.tvcolumn_settings.set_attributes(self.cell_settings,pixbuf=5)
    self.tvcolumn_settings.set_max_width(30)
    #Add delete button setup
    self.cell_delete = Gtk.CellRendererPixbuf()
    self.tvcolumn_delete = Gtk.TreeViewColumn('')
    self.treeview.append_column(self.tvcolumn_delete)
    self.tvcolumn_delete.pack_end(self.cell_delete, False)
    self.tvcolumn_delete.set_attributes(self.cell_delete,pixbuf=6)
    self.tvcolumn_delete.set_max_width(30)

    # make treeview searchable
    self.treeview.set_search_column(1)
    # Allow drag and drop reordering of rows
    self.treeview.set_reorderable(True)
    #Add treeview to base window
    self.base_area.add(self.treeview)

    #header
    self.add_conx_rows(self.conn_filter_val)
    self.show_all()

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.save_settings,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    #self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    self.apply_button.connect('clicked',self.save_settings,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    #self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def tree_item_clicked(self, treeview, event):
    pthinfo = treeview.get_path_at_pos(event.x, event.y)
    if event.button == 1: #left click
      if pthinfo != None:
        path,column,cellx,celly = pthinfo
        treeview.grab_focus()
        treeview.set_cursor(path,column,0)
        print(column.get_title())
        #update currently active display
        selection = treeview.get_selection()
        tree_model, tree_iter = selection.get_selected()
        #If selected column is delete icon then initiate delete of connection
        if tree_iter != None:
          #gathers the Connection column name and connection type in the row clicked on
          c_id = tree_model[tree_iter][2]
          c_type = tree_model[tree_iter][3]
          #checks if it is a delete or settings button click
          if column is self.tvcolumn_delete:
            self.confirm_delete('',c_id,tree_iter)
          elif column is self.tvcolumn_settings:
            self.open_settings_popup(c_id)
          elif column is self.tvcolumn_conx_button:
            #print('c_object',self.conx_obj_available[c_id].is_polling(c_id))
            if not self.conx_obj_available[c_id].is_polling(c_id):
              self.confirm_connect('button',path,c_id)
            else:
              self.confirm_disconnect('button',path,c_id)
      else:
        #unselect row in treeview
        selection = treeview.get_selection()
        selection.unselect_all()
    elif event.button == 3: #right click
      if pthinfo != None:
        path,col,cellx,celly = pthinfo
        treeview.grab_focus()
        treeview.set_cursor(path,col,0)
        rect = Gdk.Rectangle()
        rect.x = event.x
        rect.y = event.y + 10
        rect.width = rect.height = 1
        selection = treeview.get_selection()
        tree_model, tree_iter = selection.get_selected()
        popover = Gtk.Popover(width_request = 200)
        vbox = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
        if tree_iter is not None:
          #gathers the Tag name/Connection column text in the row clicked on
          c_id = tree_model[tree_iter][2]
          c_type = tree_model[tree_iter][3]
          #popover to add display
          edit_btn = Gtk.ModelButton(label="Edit", name=c_id)
          cb = lambda btn: self.open_settings_popup(c_id)
          edit_btn.connect("clicked", cb)
          vbox.pack_start(edit_btn, False, True, 10)
          delete_btn = Gtk.ModelButton(label="Delete", name=c_id)
          cb = lambda btn:self.confirm_delete('',c_id,tree_iter)
          delete_btn.connect("clicked", cb)
          vbox.pack_start(delete_btn, False, True, 10)
        popover.add(vbox)
        popover.set_position(Gtk.PositionType.RIGHT)
        popover.set_relative_to(treeview)
        popover.set_pointing_to(rect)
        popover.show_all()
        sc = popover.get_style_context()
        sc.add_class('popover-bg')
        sc.add_class('font-16')
        return
      else:
        return
    selection = treeview.get_selection()
    tree_model, tree_iter = selection.get_selected()

  def conx_connect_toggle(self, widget, path,conx_id):
    conx_params = self.get_connection_params(conx_id)
    tags = self.tags_available[conx_id]
    if not self.conx_obj_available[conx_id].is_polling(conx_id):
      self.conx_obj_available[conx_id].connect_connection(conx_id,conx_params,tags) #Sends you to connection method
      time.sleep(4.0)

      ################################Still have is polling method in connection manager to deal with
      got_connected = self.conx_obj_available[conx_id].is_polling(conx_id)
      if got_connected:
        self.liststore[path][0] = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/link_on.png', 30, 30)
      else:
        self.display_msg('Connection attempt Failed to {}'.format(conx_id))
        self.liststore[path][0] = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/link_off.png', 30, 30)
    else:
      print('disconnect')
      disc = self.conx_obj_available[conx_id].disconnect_connection(conx_id,conx_params,tags)
      if not disc:
        self.liststore[path][0] = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/link_off.png', 30, 30)
      else:
        self.display_msg('Disconnection Failed to {}'.format(conx_id))

  def save_settings(self,button,auto_close,*args):
    pass

  def add_conx_rows(self,filter,*args):
    connection_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Connect.png', 30, 30)
    connection_button = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Connect_button.png', 30, 30)
    settings_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/settings.png', 30, 30)
    delete_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Delete.png', 30, 30)
    conx_status_icon_on = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/link_on.png', 30, 30)
    conx_status_icon_off = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/link_off.png', 30, 30)
    for conx_num in self.connections_available:
      if filter == '' or filter == self.connections_available[conx_num]['id']:
        conx_status = self.get_conx_polling_status(self.connections_available[conx_num]['id'])
        if conx_status:
          conx_status_icon = conx_status_icon_on
        else:
          conx_status_icon = conx_status_icon_off          
        self.liststore.append([ conx_status_icon,
                                connection_button,
                                self.connections_available[conx_num]['id'],
                                self.connections_available[conx_num]['connection_type'],
                                self.connections_available[conx_num]['description'],
                                settings_icon,
                                delete_icon,
                                        ])
    self.show_all()

  def insert_connection_row(self,button,params,*args):
    connection_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Connect.png', 30, 30)
    connection_button = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Connect_button.png', 30, 30)
    settings_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/settings.png', 30, 30)
    delete_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Delete.png', 30, 30)
    conx_status_icon_on = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/link_on.png', 30, 30)
    conx_status_icon_off = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/link_off.png', 30, 30)
    conx_status = self.get_conx_polling_status(params['id'])
    if conx_status:
      conx_status_icon = conx_status_icon_on
    else:
      conx_status_icon = conx_status_icon_off 
    self.liststore.insert(0,[conx_status_icon,
                            connection_button,
                            params['id'],
                            params['connection_type'],
                            params['description'],
                            settings_icon,
                            delete_icon,
                                    ])
    self.show_all()
  
  def get_available_connections(self,*args):
    conx_items = ['id', 'connection_type', 'description']
    new_params = {}
    self.connections_available = {}   #Clears out old list
    self.conx_obj_available = {}   #Clears out old list
    count = 0
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      for c in conx_items:
        new_params[c] = getattr(conx_obj, c)
      self.connections_available[count] = new_params
      self.conx_obj_available[conx_id] = conx_obj   #Hold reference to all conx objects
      new_params = {}
      count += 1

  def get_available_tags(self,c_id,*args):
    self.tags_available = {}
    new_params = {}
    count = 1
    self.conx_tags = {}
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      if conx_obj:
        tag_items = conx_obj.return_tag_parameters()  #return list of tag parameters from the specific connection
        for tag_id,tag_obj in conx_obj.get('tags').items():
          for c in tag_items:
            new_params[c] = getattr(tag_obj, c)
          self.conx_tags[count] = new_params
          new_params = {}
          count += 1
        self.tags_available[conx_id]= self.conx_tags
        self.conx_tags = {}
        count = 1

  def get_conx_polling_status(self, id,*args):
    return False
    ####################method needs to be fixed for new process link
    #This method uses the stored conx objects to access polling status of connections
    """     if id in self.conx_obj_available:
      obj = self.conx_obj_available[id]
      return obj.polling
    else:
      #Connection doesn't exist
      return False """
  
  def scroll_to_bottom(self, adjust):
    max = adjust.get_upper()
    adjust.set_value(max)
    self.scroll.set_vadjustment(adjust)
    return False

  def add_style(self, wid, style):
    #style should be a list
    sc = wid.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def confirm_delete(self, button,conx_id,tree_iter,msg="Are you sure you want to delete this connection?", args=[]):
    popup = PopupConfirm(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      self.delete_row(conx_id)
      self.liststore.remove(tree_iter)
      self.get_available_connections()
      return True
    else:
      return False

  def confirm_connect(self, button,path,conx_id,msg="Start Connection?", args=[]):
    popup = PopupConfirm(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      self.conx_connect_toggle(button, path,conx_id)
    else:
      return False
    
  def confirm_disconnect(self, button,path,conx_id,msg="Disconnect?", args=[]):
    popup = PopupConfirm(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      self.conx_connect_toggle(button, path,conx_id)
    else:
      return False

  def delete_row(self,id,*args):
    ##$#conx_obj = self.app.link.get("connections").get(id)
    conx_obj = self.app.db.get("connections").get(id)
    if conx_obj != None:
      self.app.link.delete_connection(conx_obj,id)
    self.show_all()

  def create_connection(self,params,*args):
    #should be passing in description and connection_type as a dictionary
    new_conx = self.app.link.new_connection({"id": params['id'],
                            "connection_type": params['connection_type'],
                            "description": params['description']
                            })
    ##$#conx_obj = self.app.link.get("connections").get(params['id'])
    conx_obj = self.app.db.get("connections").get(params['id'])
    if conx_obj != None:
      self.app.link.save_connection(conx_obj)
      self.insert_connection_row(None,params)
      self.get_available_connections()
    else:
      #print('connection creation failed')
      self.display_msg(msg='connection creation failed')

  def display_msg(self,msg,*args):
    popup = PopupMessage(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      return True
    else:
      return False

  def update_connection(self,params,*args):
    ##$#conx_obj = self.app.link.get("connections").get(params['id'])
    conx_obj = self.app.db.get("connections").get(params['id'])
    if conx_obj != None:
      for key, val in params.items():
        if key == 'id' or key == 'description' or key == 'connection_type' or val == None:
          pass
        else:
          try:
            conx_obj.set(key,val)
          except KeyError as e:
            self.display_msg(msg="Connection Update Failed: {}, {}".format(e,key))
            #print(e,key)
      self.app.link.save_connection(conx_obj)

  def add_connection_popup(self,button,bad_name,*args):
    popup = AddConnectionPopup(self,bad_name,self.app,self.conx_type)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      #results = (popup.get_result())
      #self.check_duplicate_name(results)
      return True
    else:
      return False
  
  def get_connection_params(self,conx_id):
    ##$#conx_obj = self.app.link.get("connections").get(conx_id)
    conx_obj = self.app.db.get("connections").get(conx_id)
    if conx_obj != None:
      ##$#return self.app.link.get_connection_params(conx_obj,conx_id)
      return self.app.db.get_connection_params(conx_obj,conx_id)
  
  def open_settings_popup(self,conx_id,*args):
    params = self.get_connection_params(conx_id)
    popup = ConnectionSettingsPopup(self,params,self.app)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      #results = (popup.get_result())
      #self.update_connection(results)
      return True
    else:
      return False

  def close_popup(self, button):
    self.destroy()


class AddConnectionPopup(Gtk.Dialog):
  def __init__(self, parent,params,app,conx_type):
    Gtk.Dialog.__init__(self, '',parent, Gtk.DialogFlags.MODAL,
                        ()
                        )
    self.parent = parent
    self.app = app
    self.params = params
    self.conx_type = conx_type        #List of available connection types (Can use this to limit types of connections allowed to create)
    self.build_window()
    self.connect("response", self.on_response)
    self.result = {}

    self.content_area = self.get_content_area()
    self.dialog_window = Gtk.Box(width_request=500,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window )
    ### - Title Bar- ###
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=300)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)

    ### - Base Area- ###
    self.grid = Gtk.Grid(column_spacing=4, row_spacing=4, column_homogeneous=True, row_homogeneous=True,)
    self.dialog_window.pack_start(self.grid,1,1,1)

    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)
    self.load_settings()
    self.build_header()
    self.build_base()
    self.build_footer()
    if self.params:
      self.reload_popup(self.params)
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(200, 150)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(True)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self,*args):
    #Save Button
    self.save_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Save.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.save_button.add(image)
    sc = self.save_button.get_style_context()
    sc.add_class('ctrl-button')
    bx = Gtk.Box()
    bx.pack_end(self.save_button,0,0,0)
    #self.title_bar.pack_start(bx,0,0,0)
    self.save_button.connect('clicked',self.save_settings,False)

    #title
    title = Gtk.Label(label='Create New Connection')
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-18')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    #exit button
    self.exit_button = Gtk.Button(width_request = 20)
    self.exit_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.exit_button.add(image)
    #self.title_bar.pack_end(self.exit_button,0,0,1)
    sc = self.exit_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self,*args):
    self.pop_lbl = Gtk.Label('')
    self.add_style(self.pop_lbl,['text-red-color','font-14','font-bold'])
    self.grid.attach(self.pop_lbl,0,0,3,1)

    #Connection name entry
    lbl = Gtk.Label('Connection Name')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,1,1,1) 
    self.conx_name = Gtk.Entry(max_length = 100,width_request = 300,height_request = 30)
    self.conx_name.set_placeholder_text('Enter Connection Name')
    self.conx_name.set_alignment(0.5)
    self.add_style(self.conx_name,["entry","font-12"])
    self.conx_name.connect("notify::text-length", self.enable_new)
    self.grid.attach(self.conx_name,1,1,2,1)    

    #Connection Driver
    lbl = Gtk.Label('Connection Driver')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,2,1,1) 
    self.conx_driver = Gtk.ComboBoxText(width_request = 200,height_request = 30)#hexpand = True
    self.add_style(self.conx_driver,["font-18","list-select","font-bold"])
    val = 0
    for key in self.conx_type:
      self.conx_driver.append(str(val),key)
      val+= 1
    self.conx_driver.set_active(0)
    self.grid.attach(self.conx_driver,1,2,2,1)

   #Connection description entry
    lbl = Gtk.Label('Connection Description')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,3,1,1) 
    self.conx_descr = Gtk.Entry(max_length = 100,width_request = 300,height_request = 30)
    self.conx_descr.set_placeholder_text('Enter Connection Description')
    self.conx_descr.set_alignment(0.5)
    self.add_style(self.conx_descr,["entry","font-12"])
    self.conx_descr.connect("notify::text-length", self.enable_new)
    self.grid.attach(self.conx_descr,1,3,2,1)  

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.save_settings,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    self.apply_button.connect('clicked',self.save_settings,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    #self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def load_settings(self,*args):
    pass

  def save_settings(self,button,auto_close,*args):
    ####Get Results
    #{'id': '2', 'connection_type': 4, 'description': 'EthernetIP'}
    self.result['id'] = self.conx_name.get_text ()
    self.result['connection_type'] = self.conx_driver.get_active_text()
    self.result['description'] = self.conx_descr.get_text ()
    #####Check Results /Save
    if self.result['id'] ==  '':
      dup = True
    else:
      dup = False
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      if conx_id == self.result['id']:
        dup = True
    if dup:
      self.reload_popup(self.result)
    else:
      self.close_popup(False)
      self.parent.create_connection(self.result)
      self.parent.open_settings_popup(self.result['id'])

  def add_style(self, item,style):
    sc = item.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def enable_new(self, obj, prop):
    enable = (obj.get_property('text-length') > 0)
    if enable:
      self.add_style(obj,["entry","font-18","font-bold"])
    else:
      self.add_style(obj,["entry","font-12"])

  def close_popup(self, button):
    self.destroy()
  
  def on_response(self, widget, response_id):
    #{'id': '2', 'connection_type': 4, 'description': 'EthernetIP'}
    self.result['id'] = self.conx_name.get_text ()
    self.result['connection_type'] = self.conx_driver.get_active_text()
    self.result['description'] = self.conx_descr.get_text ()
  
  def get_result(self):
    return self.result
  
  def reload_popup(self,results):
    self.pop_lbl.set_label('Name Already Exists')
    self.conx_name.set_text('')
    val = 0
    for key in self.conx_type:
      if results['connection_type'] == key:
        self.conx_driver.set_active(val)
      val+= 1
    self.conx_descr.set_text(results['description'])


class ConnectionSettingsPopup(Gtk.Dialog):
  def __init__(self, parent,params,app):
    Gtk.Dialog.__init__(self, '',parent, Gtk.DialogFlags.MODAL,
                        ()
                        )
    self.params = params
    self.app = app
    self.build_window()
    #self.connect("response", self.on_response)
    self.result = {}

    self.content_area = self.get_content_area()
    self.dialog_window = Gtk.Box(height_request=200,width_request=500,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window )
    ### - Title Bar- ###
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=300)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)

    ### - Base Area- ###
    self.grid = Gtk.Grid(column_spacing=4, row_spacing=4, column_homogeneous=True, row_homogeneous=True,)
    self.dialog_window.pack_start(self.grid,1,1,1)

    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)

    self.load_settings()
    self.build_header()
    self.build_base()
    self.build_footer()
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(500, 200)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(True)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self,*args):
    #Save Button
    self.save_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Save.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.save_button.add(image)
    sc = self.save_button.get_style_context()
    sc.add_class('ctrl-button')
    bx = Gtk.Box()
    bx.pack_end(self.save_button,0,0,0)
    #self.title_bar.pack_start(bx,0,0,0)
    self.save_button.connect('clicked',self.save_settings,False)

    #title
    tit = self.params['connection_type']
    title = Gtk.Label(label=f'{tit} Connection')
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-18')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    #exit button
    self.exit_button = Gtk.Button(width_request = 20)
    self.exit_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.exit_button.add(image)
    #self.title_bar.pack_end(self.exit_button,0,0,1)
    sc = self.exit_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self,*args):
    row = 0

    #Connection name entry
    lbl = Gtk.Label('Connection Name')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,row,1,1) 
    self.conx_name = Gtk.Label(width_request = 300,height_request = 30)
    self.conx_name.set_text(self.params['id'])
    self.conx_name.set_alignment(0.5,0.5)
    self.add_style(self.conx_name,["label","font-18","font-bold"])
    #self.conx_name.connect("notify::text-length", self.enable_new)
    self.grid.attach(self.conx_name,1,row,2,1)   
    row+=1 

    #Connection description entry
    lbl = Gtk.Label('Connection Description')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,row,1,1) 
    self.conx_descr = Gtk.Label(width_request = 300,height_request = 30)
    self.conx_descr.set_text(self.params['description'])
    self.conx_descr.set_alignment(0.5,0.5)
    self.add_style(self.conx_descr,["label","font-18","font-bold"])
    #self.conx_descr.connect("notify::text-length", self.enable_new)
    self.grid.attach(self.conx_descr,1,row,2,1) 
    row+=1 

    #Pollrate
    if 'pollrate' in self.params.keys():    
      db_poll_rate = str(self.params['pollrate'])
      lbl = Gtk.Label('Connection Pollrate (sec)')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      but = Gtk.Button(width_request = 100)
      self.pollrate = Gtk.Label()
      self.pollrate.set_label(db_poll_rate)
      self.add_style(self.pollrate,['borderless-num-display','font-14','text-black-color'])
      but.add(self.pollrate)
      sc = but.get_style_context()
      sc.add_class('ctrl-button')
      but.connect('clicked',self.open_numpad,self.pollrate,{'min':0.001,'max':10000.0,'type':float,'polarity':True,'name':'Pollrate'})
      self.grid.attach(but,1,row,2,1)
      row+=1 

    #Auto Connect
    if 'auto_connect' in self.params.keys(): 
      db_auto_connect = str(self.params['auto_connect'])
      lbl = Gtk.Label('Auto Connect on Start')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      bx = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL, halign = Gtk.Align.CENTER)
      p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale(os.path.join(PUBLIC_DIR, 'images/Check.png'), 20, -1, True)
      image = Gtk.Image(pixbuf=p_buf)
      wid =CheckBoxWidget(30,30,image,db_auto_connect)
      self.auto_connect = wid.return_self()
      bx.pack_start(self.auto_connect,0,0,0)
      self.grid.attach(bx,1,row,2,1)
      row+=1

    #Connection host entry
    if 'host' in self.params.keys(): 
      db_host = str(self.params['host'])
      lbl = Gtk.Label('Connection Host')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      self.conx_host = Gtk.Entry(max_length = 100,width_request = 300,height_request = 30)
      self.conx_host.set_alignment(0.5)
      self.add_style(self.conx_host,["entry","font-18","font-bold"])
      self.conx_host.set_text(db_host)
      self.grid.attach(self.conx_host,1,row,2,1)
      row+=1 

    #Port
    if 'port' in self.params.keys(): 
      db_port = str(self.params['port'])
      lbl = Gtk.Label('Connection Port')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      but = Gtk.Button(width_request = 100)
      self.conx_port = Gtk.Label()
      self.conx_port.set_label(db_port)
      self.add_style(self.conx_port,['borderless-num-display','font-14','text-black-color'])
      but.add(self.conx_port)
      sc = but.get_style_context()
      sc.add_class('ctrl-button')
      but.connect('clicked',self.open_numpad,self.conx_port,{'min':0,'max':65536,'type':int,'polarity':True,'name':'Port Number'})
      self.grid.attach(but,1,row,2,1)
      row+=1 

    #Station ID
    if 'station_id' in self.params.keys(): 
      db_station_id = str(self.params['station_id'])
      lbl = Gtk.Label('Station ID')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      but = Gtk.Button(width_request = 100)
      self.station_id = Gtk.Label()
      self.station_id.set_label(db_station_id)
      self.add_style(self.station_id,['borderless-num-display','font-14','text-black-color'])
      but.add(self.station_id)
      sc = but.get_style_context()
      sc.add_class('ctrl-button')
      but.connect('clicked',self.open_numpad,self.station_id,{'min':0,'max':255,'type':int,'polarity':True,'name':'Station ID'})
      self.grid.attach(but,1,row,2,1)
      row+=1 

    #Connection Baudrate
    if 'baudrate' in self.params.keys():
      db_baudrate = str(self.params['baudrate'])
      lbl = Gtk.Label('Baudrate')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      self.baudrate = Gtk.ComboBoxText(width_request = 200,height_request = 30)#hexpand = True
      self.add_style(self.baudrate,["font-18","list-select","font-bold"])
      br = None
      val = 0
      b_rates = ['1200','2400','4800','9600','19200','38400','57600','115200']
      for item in b_rates:
        self.baudrate.append(str(val),item)
        if item == db_baudrate:
          br = val
        val+= 1
      if br:
        self.baudrate.set_active(br)
      else:
        self.baudrate.set_active(0)
      self.grid.attach(self.baudrate,1,row,2,1)
      row+=1 

    #Timeout
    if 'timeout' in self.params.keys(): 
      db_timeout = str(self.params['timeout'])
      lbl = Gtk.Label('Timeout (sec)')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      but = Gtk.Button(width_request = 100)
      self.timeout = Gtk.Label()
      self.timeout.set_label(db_timeout)
      self.add_style(self.timeout,['borderless-num-display','font-14','text-black-color'])
      but.add(self.timeout)
      sc = but.get_style_context()
      sc.add_class('ctrl-button')
      but.connect('clicked',self.open_numpad,self.timeout,{'min':0,'max':100,'type':int,'polarity':True,'name':'Timeout (sec)'})
      self.grid.attach(but,1,row,2,1)
      row+=1 

    #Stop Bit
    if 'stop_bit' in self.params.keys():
      db_stop_bit = str(self.params['stop_bit'])
      lbl = Gtk.Label('Stop bits')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      self.stop_bit = Gtk.ComboBoxText(width_request = 200,height_request = 30)#hexpand = True
      self.add_style(self.stop_bit,["font-18","list-select","font-bold"])
      sb = None
      val = 0
      s_bits = ['0','1','2']
      for item in s_bits:
        self.stop_bit.append(str(val),item)
        if item == db_stop_bit:
          sb = val
        val+= 1
      if sb:
        self.stop_bit.set_active(sb)
      else:
        self.stop_bit.set_active(0)
      self.grid.attach(self.stop_bit,1,row,2,1)
      row+=1 

    #parity
    if 'parity' in self.params.keys():
      db_parity = str(self.params['parity'])
      lbl = Gtk.Label('Parity')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      self.parity = Gtk.ComboBoxText(width_request = 200,height_request = 30)#hexpand = True
      self.add_style(self.parity,["font-18","list-select","font-bold"])
      val = 0
      p_type = ['N','O','E']
      for item in p_type:
        self.parity.append(str(val),item)
        if item == db_parity:
          p = val
        val+= 1
      if p:
        self.parity.set_active(p)
      else:
        self.parity.set_active(0)
      self.grid.attach(self.parity,1,row,2,1)
      row+=1 

    #byte_size
    if 'byte_size' in self.params.keys():
      db_byte_size = str(self.params['byte_size'])
      lbl = Gtk.Label('Byte Size')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      self.byte_size = Gtk.ComboBoxText(width_request = 200,height_request = 30)#hexpand = True
      self.add_style(self.byte_size,["font-18","list-select","font-bold"])
      bs = None
      val = 0
      b_size = ['8','7']
      for item in b_size:
        self.byte_size.append(str(val),item)
        if item == db_byte_size:
          bs = val
        val+= 1
      if bs:
        self.byte_size.set_active(bs)
      else:
        self.byte_size.set_active(0)
      self.grid.attach(self.byte_size,1,row,2,1)
      row+=1 

    #Retries
    if 'retries' in self.params.keys(): 
      db_retries = str(self.params['retries'])
      lbl = Gtk.Label('Retries')
      self.add_style(lbl,["Label","font-16",'font-bold'])
      self.grid.attach(lbl,0,row,1,1) 
      but = Gtk.Button(width_request = 100)
      self.retries = Gtk.Label()
      self.retries.set_label(db_retries)
      self.add_style(self.retries,['borderless-num-display','font-14','text-black-color'])
      but.add(self.retries)
      sc = but.get_style_context()
      sc.add_class('ctrl-button')
      but.connect('clicked',self.open_numpad,self.retries,{'min':0,'max':10,'type':int,'polarity':True,'name':'Retries'})
      self.grid.attach(but,1,row,2,1)
      row+=1 
    
  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.save_settings,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    self.apply_button.connect('clicked',self.save_settings,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def load_settings(self,*args):
    pass

  def save_settings(self,button,auto_close,*args):
    self.result['id'] = self.conx_name.get_text()
    self.result['connection_type'] = self.params['connection_type']
    self.result['description'] = self.conx_descr.get_text ()
    if 'pollrate' in self.params.keys():
      self.result['pollrate'] = self.pollrate.get_text()
    else:
      self.result['pollrate'] = None
    if 'auto_connect' in self.params.keys():
      self.result['auto_connect'] = self.auto_connect.get_active()
    else:
      self.result['auto_connect'] = None
    if 'host' in self.params.keys():
      self.result['host'] = self.conx_host.get_text()
    else:
      self.result['host'] = None
    if 'port' in self.params.keys():
      self.result['port'] = self.conx_port.get_text()
    else:
      self.result['port'] = None
    if 'station_id' in self.params.keys():
      self.result['station_id'] = self.station_id.get_text()
    else:
      self.result['station_id'] = None
    if 'baudrate' in self.params.keys():
      self.result['baudrate'] = self.baudrate.get_active_text()
    else:
      self.result['baudrate'] = None
    if 'timeout' in self.params.keys():
      self.result['timeout'] = self.timeout.get_text()
    else:
      self.result['timeout'] = None
    if 'stop_bit' in self.params.keys():
      self.result['stop_bit'] = self.stop_bit.get_active_text()
    else:
      self.result['stop_bit'] = None
    if 'parity' in self.params.keys():
      self.result['parity'] = self.parity.get_active_text()
    else:
      self.result['parity'] = None
    if 'byte_size' in self.params.keys():
      self.result['byte_size'] = self.byte_size.get_active_text()
    else:
      self.result['byte_size'] = None
    if 'retries' in self.params.keys():
      self.result['retries'] = self.retries.get_text()
    else:
      self.result['retries'] = None
    ####Save values
    ##$#conx_obj = self.app.link.get("connections").get(self.result['id'])
    conx_obj = self.app.db.get("connections").get(self.result['id'])
    if conx_obj != None:
      for key, val in self.result.items():
        if key == 'id' or key == 'description' or key == 'connection_type' or val == None:
          pass
        else:
          try:
            conx_obj.set(key,val)
          except KeyError as e:
            print(e,key)
      self.app.link.save_connection(conx_obj)
      if auto_close:
        self.close_popup(False)

  def add_style(self, item,style):
    sc = item.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def enable_new(self, obj, prop):
    enable = (obj.get_property('text-length') > 0)
    if enable:
      self.add_style(obj,["entry","font-18","font-bold"])
    else:
      self.add_style(obj,["entry","font-12"])

  def close_popup(self, button):
    self.destroy()
  
  """ def on_response(self, widget, response_id):
    self.result['id'] = self.conx_name.get_text()
    self.result['connection_type'] = self.params['connection_type']
    self.result['description'] = self.conx_descr.get_text ()
    if 'pollrate' in self.params.keys():
      self.result['pollrate'] = self.pollrate.get_text()
    else:
      self.result['pollrate'] = None
    if 'auto_connect' in self.params.keys():
      self.result['auto_connect'] = self.auto_connect.get_active()
    else:
      self.result['auto_connect'] = None
    if 'host' in self.params.keys():
      self.result['host'] = self.conx_host.get_text()
    else:
      self.result['host'] = None
    if 'port' in self.params.keys():
      self.result['port'] = self.conx_port.get_text()
    else:
      self.result['port'] = None
    if 'station_id' in self.params.keys():
      self.result['station_id'] = self.station_id.get_text()
    else:
      self.result['station_id'] = None
    if 'baudrate' in self.params.keys():
      self.result['baudrate'] = self.baudrate.get_active_text()
    else:
      self.result['baudrate'] = None
    if 'timeout' in self.params.keys():
      self.result['timeout'] = self.timeout.get_text()
    else:
      self.result['timeout'] = None
    if 'stop_bit' in self.params.keys():
      self.result['stop_bit'] = self.stop_bit.get_active_text()
    else:
      self.result['stop_bit'] = None
    if 'parity' in self.params.keys():
      self.result['parity'] = self.parity.get_active_text()
    else:
      self.result['parity'] = None
    if 'byte_size' in self.params.keys():
      self.result['byte_size'] = self.byte_size.get_active_text()
    else:
      self.result['byte_size'] = None
    if 'retries' in self.params.keys():
      self.result['retries'] = self.retries.get_text()
    else:
      self.result['retries'] = None """
  
  def get_result(self):
    return self.result

  def open_numpad(self,button,widget_obj,params,*args):
    numpad = ValueEnter(self,widget_obj,params)
    response = numpad.run()
    if response == Gtk.ResponseType.NO:
      pass
    else:
      pass
    numpad.destroy()
  

class ValueEnter(Gtk.Dialog):
  #Need to add check for value exceeding min,max range based on type
  def __init__(self, parent,obj,params):
    super().__init__(flags=0) 

    self.params = params
    self.widget_obj = obj
    self.first_key_pressed = False #the user hasn't typed anything yet
    self.initial_value = 0
    self.build_window()

    self.content_area = self.get_content_area()
    self.dialog_window = Gtk.Box(width_request=600,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window )
    ### - Title Bar- ###
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)

    ### - Base Area- ###
    self.grid = Gtk.Grid(column_spacing=4, row_spacing=4, column_homogeneous=True, row_homogeneous=True,)
    self.dialog_window.pack_start(self.grid,1,1,1)

    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)

    self.build_header()
    self.build_base()
    self.build_footer()
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(600, 400)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(True)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self,*args):
    #Save Button
    self.save_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Save.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.save_button.add(image)
    sc = self.save_button.get_style_context()
    sc.add_class('ctrl-button')
    bx = Gtk.Box()
    bx.pack_end(self.save_button,0,0,0)
    #self.title_bar.pack_start(bx,0,0,0)
    self.save_button.connect('clicked',self.save_settings,False)

    #title
    title = Gtk.Label(label='Numpad')
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-18')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    #exit button
    self.exit_button = Gtk.Button(width_request = 20)
    self.exit_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.exit_button.add(image)
    #self.title_bar.pack_end(self.exit_button,0,0,1)
    sc = self.exit_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self,*args):
    if 'name' in self.params:
      lbl = self.params['name']
    else:
      lbl = 'Numpad'
    pop_lbl = Gtk.Label("{}".format(lbl))
    self.add_style(pop_lbl,['borderless-num-display','font-14','text-black-color'])
    self.grid.attach(pop_lbl,0,0,2,1)

    try:
      if isinstance(self.widget_obj, Gtk.Label):
        value = self.params['type'](self.widget_obj.get_label())
      if isinstance(self.widget_obj, Gtk.Entry):
        value = self.params['type'](self.widget_obj.get_text())
      if isinstance(self.widget_obj,Gtk.CellRendererText):
        value = self.params['value']
      self.initial_value = value
    except ValueError:
      value = 0

    #value = 0
    self.val_label = Gtk.Label(str(value))
    self.add_style(self.val_label,['numpad-display','font-16'])
    self.grid.attach(self.val_label,2,0,1,1)
    min_str = "-"+chr(0x221e) if type(self.params['min']) == type(None) else self.params['min']
    max_str = chr(0x221e) if type(self.params['max']) == type(None) else self.params['max']
    min_max_lbl = Gtk.Label(u"({} ~ {})".format(min_str, max_str))
    self.add_style(min_max_lbl,['font-14'])
    self.grid.attach(min_max_lbl,3,0,1,1)
    key = []
    for k in range(10):
      b = Gtk.Button(str(k), can_focus=False, can_default=False)
      #b.get_style_context().add_class("keypad_key")
      b.connect("clicked", self.btn_pressed)
      key.append(b)
      self.add_style(b,['numpad-bg','keypad_key'])
    self.grid.attach(key[7],0,2,1,1)
    self.grid.attach(key[8],1,2,1,1)
    self.grid.attach(key[9],2,2,1,1)
  
    self.grid.attach(key[4],0,3,1,1)
    self.grid.attach(key[5],1,3,1,1)
    self.grid.attach(key[6],2,3,1,1)

    self.grid.attach(key[1],0,4,1,1)
    self.grid.attach(key[2],1,4,1,1)
    self.grid.attach(key[3],2,4,1,1)

    self.grid.attach(key[0],0,5,2,1)

    period_key = Gtk.Button(".", can_focus=False, can_default=False)
    period_key.connect("clicked", self.add_period)
    self.add_style(period_key,['numpad-bg','keypad_key'])
    if self.params['type'] == float:
      self.grid.attach(period_key,2,5,1,1)

    PM_key = Gtk.Button("+/-")
    PM_key.connect("clicked", self.add_plus_minus)
    self.add_style(PM_key,['numpad-bg','keypad_key'])
    if self.params['polarity']:
      self.grid.attach(PM_key,3,5,1,1)
    
    clear_key = Gtk.Button("CLEAR", can_focus=False, can_default=False)
    clear_key.connect("clicked", self.init_val)
    self.add_style(clear_key,['numpad-cmd-bg','keypad_enter'])
    self.grid.attach(clear_key,3,2,1,1)

    delete_key = Gtk.Button("DEL", can_focus=False, can_default=False)
    delete_key.connect("clicked", self.del_num)
    self.add_style(delete_key,['numpad-cmd-bg','keypad_enter'])
    self.grid.attach(delete_key,3,3,1,1)

    enter_key = Gtk.Button("ENTER", can_focus=False, can_default=False)
    enter_key.connect("clicked", self.accept_val)
    self.add_style(enter_key,['numpad-cmd-bg','keypad_enter'])
    self.grid.attach(enter_key,3,4,1,1)


    self.signals = []
    self.signals.append(self.connect('key-release-event', self.key_pressed))
    self.show_all()

    #Add style to dialog buttons
    a = self.get_action_area()
    b = a.get_children()
    for but in b:
      self.add_style(but,['dialog-buttons','font-16'])

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.save_settings,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    #self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    self.apply_button.connect('clicked',self.save_settings,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    #self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def key_pressed(self, popup, key_event):
    if key_event.get_keycode().keycode == 13:#Enter
      self.accept_val(None)
    if key_event.get_keycode().keycode == 8:#Backspace
      self.first_key_pressed = True #they want to use the number in there
      self.del_num(None)
    elif key_event.string == "-" or key_event.string == "+":
      self.add_plus_minus(None)
    elif key_event.string in "0123456789" and len(key_event.string)==1:
      self.update_val(key_event.string)
    elif key_event.string == ".":
      self.add_period(None)
    else:
      pass

  def add_style(self, item,style):
    sc = item.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def load_settings(self,*args):
    pass

  def save_settings(self,button,auto_close,*args):
    pass

  def btn_pressed(self, key):
    num = int(key.get_label())
    self.update_val(num)

  def update_val(self, num):
    if not self.first_key_pressed:
      val = str(num)
      self.first_key_pressed = True
    else:
      old_val = self.val_label.get_text()
      if old_val == '0':
        val = str(num)
      else:
        val = old_val+str(num)
    if self.check_limits(val):
      self.val_label.set_text(val)
  
  def check_limits(self, val):
    try:
      val = float(val)
    except ValueError:
      return False
    if not type(self.params['min']) == type(None) and val < self.params['min']:
      return False    
    if not type(self.params['max']) == type(None) and val > self.params['max']:
      return False
    return True
  
  def init_val(self, key):
    self.val_label.set_text('')
  
  def del_num(self,*args):
    val = self.val_label.get_text()
    val = (val)[:-1]
    if not val:
      val = '0'
    self.val_label.set_text(val)

  def add_period(self,*args):
    if not self.first_key_pressed:
      self.update_val("0")
    val = self.val_label.get_text()
    if "." not in val:
      val = val+"."
      self.val_label.set_text(val)

  def add_plus_minus(self,*args):
    val = self.val_label.get_text()
    if "-" not in val:
      val = "-"+val
    else:
      val = val.replace('-',"")
    if self.check_limits(val):
      self.val_label.set_text(val)

  def cleanup(self):
    for signal in self.signals:
      self.disconnect(signal)

  def accept_val(self, key):
    self.cleanup()
    try:
      if isinstance(self.widget_obj, Gtk.Label):
        self.widget_obj.set_label((self.val_label.get_text()))
        value = self.val_label.get_text()
      if isinstance(self.widget_obj, Gtk.Entry):
        self.widget_obj.set_text(str(self.val_label.get_text()))
        value = self.val_label.get_text()
      if isinstance(self.widget_obj,Gtk.CellRendererText):
        self.params['ref_val'][2] = (str(self.val_label.get_text()))
        value = self.val_label.get_text()
      else:
        value = 0
    except ValueError:
      value = 0
      pass
    self.destroy()
    if str(self.initial_value) != str(value):
      pass
    else:
      pass

  def close_popup(self, button):
    self.destroy()


class PopupConfirm(Gtk.Dialog):
  def __init__(self, parent, msg='Do you really want to do this?'):
      Gtk.Dialog.__init__(self, "Confirm?", parent, Gtk.DialogFlags.MODAL,
                          (Gtk.STOCK_YES, Gtk.ResponseType.YES,
                            Gtk.STOCK_NO, Gtk.ResponseType.NO)
                          )
      self.set_default_size(300, 200)
      self.set_border_width(10)
      sc = self.get_style_context()
      sc.add_class("dialog-border")
      self.set_keep_above(True)
      self.set_decorated(False)
      box = Gtk.Box()
      box.set_spacing(10)
      box.set_orientation(Gtk.Orientation.VERTICAL)
      c = self.get_content_area()
      c.add(box)
      box.pack_start(Gtk.Image(stock=Gtk.STOCK_DIALOG_WARNING), 0, 0, 0)
      confirm_msg = Gtk.Label(msg + '\n\n')
      sc = confirm_msg.get_style_context()
      sc.add_class('borderless-num-display')
      sc.add_class('text-black-color')
      sc.add_class('font-20')
      box.pack_start(confirm_msg, 0, 0, 0)
      sep = Gtk.Label(height_request=3)
      c.pack_start(sep,1,1,1)
      self.show_all()
      #Add style to dialog buttons
      a = self.get_action_area()
      b = a.get_children()
      for but in b:
        sc = but.get_style_context()
        sc.add_class("dialog-buttons")
        sc.add_class("font-16")


class PopupMessage(Gtk.Dialog):
  def __init__(self, parent, msg='No Messaged Generated'):
      Gtk.Dialog.__init__(self, "Confirm?", parent, Gtk.DialogFlags.MODAL,
                          (Gtk.STOCK_OK, Gtk.ResponseType.YES)
                          )
      self.set_default_size(300, 200)
      self.set_border_width(10)
      sc = self.get_style_context()
      sc.add_class("dialog-border")
      self.set_keep_above(True)
      self.set_decorated(False)
      box = Gtk.Box()
      box.set_spacing(10)
      box.set_orientation(Gtk.Orientation.VERTICAL)
      c = self.get_content_area()
      c.add(box)
      box.pack_start(Gtk.Image(stock=Gtk.STOCK_DIALOG_WARNING), 0, 0, 0)
      confirm_msg = Gtk.Label(msg + '\n\n')
      sc = confirm_msg.get_style_context()
      sc.add_class('borderless-num-display')
      sc.add_class('text-black-color')
      sc.add_class('font-20')
      box.pack_start(confirm_msg, 0, 0, 0)
      sep = Gtk.Label(height_request=3)
      c.pack_start(sep,1,1,1)
      self.show_all()
      #Add style to dialog buttons
      a = self.get_action_area()
      b = a.get_children()
      for but in b:
        sc = but.get_style_context()
        sc.add_class("dialog-buttons")
        sc.add_class("font-16")


class ChartSettingsPopup(Gtk.Dialog):
  def __init__(self, app,chart):
    Gtk.Dialog.__init__(self, '',None, Gtk.DialogFlags.MODAL,
                        ()
                        )
    self.chart = chart
    self.c_id = self.chart.db_id
    self.app = app
    self.db_session = self.app.settings_db.session
    self.db_model = self.app.settings_db.models['chart']
    self.Tbl = self.db_model
    self.build_window()
    self.bg_color = [1.0,1.0,1.0,1.0] #default to white
    self.grid_color = [1.0,1.0,1.0,1.0] #default to white
    self.marker1_color = [1.0,0.0,0.0,1.0] #default to red
    self.marker2_color = [0.0,1.0,0.0,1.0] #default to blue
    self.h_grids = 2
    self.v_grids = 2
    self.marker1_width = 1
    self.marker2_width = 1


    self.content_area = self.get_content_area()
    self.dialog_window = Gtk.Box(width_request=300,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window )
    ### - Title Bar- ###
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=300)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)

    ### - Base Area- ###
    self.grid = Gtk.Grid(column_spacing=3, row_spacing=4, column_homogeneous=False, row_homogeneous=True,)
    self.dialog_window.pack_start(self.grid,1,1,1)

    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)
    self.load_settings()
    self.build_header()
    self.build_base()
    self.build_footer()
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(300, 300)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(False)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self,*args):
    #Save Button
    self.save_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Save.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.save_button.add(image)
    sc = self.save_button.get_style_context()
    sc.add_class('ctrl-button')
    bx = Gtk.Box()
    bx.pack_end(self.save_button,0,0,0)
    #self.title_bar.pack_start(bx,0,0,0)
    self.save_button.connect('clicked',self.save_settings)

    #title
    title = Gtk.Label(label='Chart Settings')
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-18')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    #exit button
    self.exit_button = Gtk.Button(width_request = 20)
    self.exit_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.exit_button.add(image)
    #self.title_bar.pack_end(self.exit_button,0,0,1)
    sc = self.exit_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self,*args):
    #Chart Select
    lbl = Gtk.Label('Chart Number')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,1,1,1) 
    self.chart_select = Gtk.ComboBoxText(width_request = 200,height_request = 30)#hexpand = True
    self.add_style(self.chart_select,["font-18","list-select","font-bold"])
    selections = []
    for num in range(16):
      self.chart_select.append(str(num+1),str(num+1))
    self.chart_select.set_active((self.c_id-1))
    self.grid.attach(self.chart_select,1,1,1,1)
    self.chart_select.connect("changed", self.new_chart_selected)

    #color
    lbl = Gtk.Label('Background Color')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,2,1,1) 
    rgbcolor = Gdk.RGBA()
    rgbcolor.red = float(self.bg_color[0])
    rgbcolor.green = float(self.bg_color[1])
    rgbcolor.blue = float(self.bg_color[2])
    rgbcolor.alpha = float(self.bg_color[3])
    bx = Gtk.Box()
    self.color_button = Gtk.ColorButton(width_request = 50)
    self.color_button.set_rgba (rgbcolor)
    sc = self.color_button.get_style_context()
    sc.add_class('ctrl-button')
    bx.set_center_widget(self.color_button)
    self.grid.attach(bx,1,2,1,1)
    #self.color_button.connect('color-set',self.row_changed)

    #Horizontal Grid LInes
    lbl = Gtk.Label('Horizontal Grid Lines')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,3,1,1) 
    but = Gtk.Button(width_request = 20)
    self.hor_grid = Gtk.Label()
    self.hor_grid.set_label(str(self.h_grids))
    self.add_style(self.hor_grid,['borderless-num-display','font-14','text-black-color'])
    but.add(self.hor_grid)
    sc = but.get_style_context()
    sc.add_class('ctrl-button')
    but.connect('clicked',self.open_numpad,self.hor_grid,{'min':0,'max':8,'type':int,'polarity':False,'name':'Horizontal Grid Lines'})
    self.grid.attach(but,1,3,1,1)
    #but.connect('clicked',self.row_changed)

    #Vertical Grid Lines
    lbl = Gtk.Label('Vertical Grid Lines')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,4,1,1) 
    but = Gtk.Button(width_request = 20)
    self.vert_grid = Gtk.Label()
    self.vert_grid.set_label(str(self.v_grids))
    self.add_style(self.vert_grid,['borderless-num-display','font-14','text-black-color'])
    but.add(self.vert_grid)
    sc = but.get_style_context()
    sc.add_class('ctrl-button')
    but.connect('clicked',self.open_numpad,self.vert_grid,{'min':0,'max':8,'type':int,'polarity':False,'name':'Vertical Grid Lines'})
    self.grid.attach(but,1,4,1,1)
    #but.connect('clicked',self.row_changed)

    #grid color
    lbl = Gtk.Label('Grid Line Color')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,5,1,1) 
    rgbcolor = Gdk.RGBA()
    rgbcolor.red = float(self.grid_color[0])
    rgbcolor.green = float(self.grid_color[1])
    rgbcolor.blue = float(self.grid_color[2])
    rgbcolor.alpha = float(self.grid_color[3])
    bx = Gtk.Box()
    self.grid_color_button = Gtk.ColorButton(width_request = 50)
    self.grid_color_button.set_rgba (rgbcolor)
    sc = self.grid_color_button.get_style_context()
    sc.add_class('ctrl-button')
    bx.set_center_widget(self.grid_color_button)
    self.grid.attach(bx,1,5,1,1)

    #Marker 1
    lbl = Gtk.Label('Chart Marker 1 (Width/Color)')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,6,1,1) 
    but = Gtk.Button(width_request = 20)
    self.marker1_width_button = Gtk.Label()
    self.marker1_width_button.set_label(str(self.marker1_width))
    self.add_style(self.marker1_width_button,['borderless-num-display','font-14','text-black-color'])
    but.add(self.marker1_width_button)
    sc = but.get_style_context()
    sc.add_class('ctrl-button')
    but.connect('clicked',self.open_numpad,self.marker1_width_button,{'min':0,'max':8,'type':int,'polarity':False,'name':'Marker 1 Width'})
    self.grid.attach(but,1,6,1,1)
    #but.connect('clicked',self.row_changed)

    rgbcolor = Gdk.RGBA()
    rgbcolor.red = float(self.marker1_color[0])
    rgbcolor.green = float(self.marker1_color[1])
    rgbcolor.blue = float(self.marker1_color[2])
    rgbcolor.alpha = float(self.marker1_color[3])
    bx = Gtk.Box()
    self.marker1_color_button = Gtk.ColorButton(width_request = 50)
    self.marker1_color_button.set_rgba (rgbcolor)
    sc = self.marker1_color_button.get_style_context()
    sc.add_class('ctrl-button')
    bx.set_center_widget(self.marker1_color_button)
    self.grid.attach(bx,2,6,1,1)

    #Marker 2
    lbl = Gtk.Label('Chart Marker 2 (Width/Color)')
    self.add_style(lbl,["Label","font-16",'font-bold'])
    self.grid.attach(lbl,0,7,1,1) 
    but = Gtk.Button(width_request = 20)
    self.marker2_width_button = Gtk.Label()
    self.marker2_width_button.set_label(str(self.marker2_width))
    self.add_style(self.marker2_width_button,['borderless-num-display','font-14','text-black-color'])
    but.add(self.marker2_width_button)
    sc = but.get_style_context()
    sc.add_class('ctrl-button')
    but.connect('clicked',self.open_numpad,self.marker2_width_button,{'min':0,'max':8,'type':int,'polarity':False,'name':'Marker 2 Width'})
    self.grid.attach(but,1,7,1,1)
    #but.connect('clicked',self.row_changed)

    rgbcolor = Gdk.RGBA()
    rgbcolor.red = float(self.marker2_color[0])
    rgbcolor.green = float(self.marker2_color[1])
    rgbcolor.blue = float(self.marker2_color[2])
    rgbcolor.alpha = float(self.marker2_color[3])
    bx = Gtk.Box()
    self.marker2_color_button = Gtk.ColorButton(width_request = 50)
    self.marker2_color_button.set_rgba (rgbcolor)
    sc = self.marker2_color_button.get_style_context()
    sc.add_class('ctrl-button')
    bx.set_center_widget(self.marker2_color_button)
    self.grid.attach(bx,2,7,1,1)

    sep = Gtk.Label(height_request=3)
    self.dialog_window.pack_start(sep,1,1,1)

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.save_settings,self.c_id,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    self.apply_button.connect('clicked',self.save_settings,self.c_id,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def load_settings(self,*args):
    settings = self.db_session.query(self.Tbl).filter(self.Tbl.id == int(self.c_id)).first()
    if settings:
      self.bg_color = json.loads(settings.bg_color) #rgb in json
      self.h_grids = settings.h_grids
      self.v_grids = settings.v_grids
      self.grid_color = json.loads(settings.grid_color) #rgb in json
      self.marker1_width = settings.marker1_width
      self.marker1_color = json.loads(settings.marker1_color) #rgb in json 
      self.marker2_width = settings.marker2_width
      self.marker2_color = json.loads(settings.marker2_color) #rgb in json 
    else:
      print("Chart Not Found")
      #using default values
  def confirm(self, button,pen_id,msg="Are you sure you want to delete this pen?", args=[]):
    popup = PopupConfirm(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      self.delete_row(pen_id)
      return True
    else:
      return False

  def new_chart_selected(self, but,*args):
    val = self.chart_select.get_active_text()
    self.c_id = int(val)
    self.load_settings()
    self.vert_grid.set_label(str(self.v_grids))
    self.hor_grid.set_label(str(self.h_grids))
    bg_rgbcolor = Gdk.RGBA()
    bg_rgbcolor.red = float(self.bg_color[0])
    bg_rgbcolor.green = float(self.bg_color[1])
    bg_rgbcolor.blue = float(self.bg_color[2])
    bg_rgbcolor.alpha = float(self.bg_color[3])
    self.color_button.set_rgba (bg_rgbcolor)
    grid_rgbcolor = Gdk.RGBA()
    grid_rgbcolor.red = float(self.grid_color[0])
    grid_rgbcolor.green = float(self.grid_color[1])
    grid_rgbcolor.blue = float(self.grid_color[2])
    grid_rgbcolor.alpha = float(self.grid_color[3])
    self.grid_color_button.set_rgba (grid_rgbcolor)
    marker1_rgbcolor = Gdk.RGBA()
    marker1_rgbcolor.red = float(self.marker1_color[0])
    marker1_rgbcolor.green = float(self.marker1_color[1])
    marker1_rgbcolor.blue = float(self.marker1_color[2])
    marker1_rgbcolor.alpha = float(self.marker1_color[3])
    self.marker1_color_button.set_rgba (marker1_rgbcolor)
    marker2_rgbcolor = Gdk.RGBA()
    marker2_rgbcolor.red = float(self.marker2_color[0])
    marker2_rgbcolor.green = float(self.marker2_color[1])
    marker2_rgbcolor.blue = float(self.marker2_color[2])
    marker2_rgbcolor.alpha = float(self.marker2_color[3])
    self.marker2_color_button.set_rgba (marker2_rgbcolor)

  def save_settings(self,but,chart_id,auto_close,*args):
    self.c_id = int(self.chart_select.get_active_text())
    bg_color = self.color_button.get_rgba()
    bg_color_list = []
    for c in bg_color:
      bg_color_list.append(c)

    grid_color = self.grid_color_button.get_rgba()
    grid_color_list = []
    for c in grid_color:
      grid_color_list.append(c)

    marker1_color = self.marker1_color_button.get_rgba()
    marker1_color_list = []
    for c in marker1_color:
      marker1_color_list.append(c)

    marker2_color = self.marker2_color_button.get_rgba()
    marker2_color_list = []
    for c in marker2_color:
      marker2_color_list.append(c)

    settings = self.db_session.query(self.Tbl).filter(self.Tbl.id == int(self.c_id)).first()
    if settings:
      settings.bg_color =  json.dumps(bg_color_list)
      settings.h_grids =  int(self.hor_grid.get_label())
      settings.v_grids = int(self.vert_grid.get_label())
      settings.grid_color =  json.dumps(grid_color_list)
      settings.marker1_width = int(self.marker1_width_button.get_label())
      settings.marker1_color =  json.dumps(marker1_color_list)
      settings.marker2_width = int(self.marker2_width_button.get_label())
      settings.marker2_color =  json.dumps(marker2_color_list)
      self.db_session.commit()
    self.app.charts[self.c_id].reload_chart()
    if auto_close:
      self.close_popup()

  def remove_widgets(self,*args):
    grid = self.grid.get_children()
    for widgets in grid:
      self.grid.remove(widgets)

  def add_style(self, item,style):
    sc = item.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def open_numpad(self,button,widget_obj,params,*args):
    numpad = ValueEnter(self,widget_obj,params)
    response = numpad.run()
    if response == Gtk.ResponseType.NO:
      pass
    else:
      pass
    numpad.destroy()

  def close_popup(self, *args):
    self.destroy()


class TimeSpanPopup(Gtk.Dialog):
  def __init__(self, app,chart):
    Gtk.Dialog.__init__(self, '',None, Gtk.DialogFlags.MODAL,
                        ()
                        )
    self.app = app
    self.chart = chart
    self.c_id = self.chart.db_id
    self.db_session = self.app.settings_db.session
    self.db_model = self.app.settings_db.models['chart']
    self.Tbl = self.db_model
    self.build_window()
    self.content_area = self.get_content_area()
    self.dialog_window = Gtk.Box(width_request=300,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window )
    ### - Title Bar- ###
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=300)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)

    ### - Base Area- ###
    self.base_area = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
    self.dialog_window.pack_start(self.base_area, 0, 0, 0)

    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)
    self.build_header()
    self.build_base()
    self.build_footer()
    self.load_settings()
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(300, 300)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(False)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self,*args):
    #Save Button
    self.save_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Save.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.save_button.add(image)
    sc = self.save_button.get_style_context()
    sc.add_class('ctrl-button')
    bx = Gtk.Box()
    bx.pack_end(self.save_button,0,0,0)
    #self.title_bar.pack_start(bx,0,0,0)
    self.save_button.connect('clicked',self.save_settings,self.c_id)

    #title
    title = Gtk.Label(label='Chart Time')
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-22')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    #exit button
    self.exit_button = Gtk.Button(width_request = 20)
    self.exit_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.exit_button.add(image)
    #self.title_bar.pack_end(self.exit_button,0,0,1)
    sc = self.exit_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self,*args):
    #Date / Time Picker

    hbox = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
    vbox = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)

    #Timespan
    bx = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
    lbl = Gtk.Label('Chart {} Timespan (min)'.format(self.c_id))
    self.add_style(lbl,["Label","font-18",'font-bold'])
    bx.pack_start(lbl,1,1,1)
    but = Gtk.Button(width_request = 200)
    self.timespan = Gtk.Label()
    self.add_style(self.timespan,['borderless-num-display','font-18','text-black-color'])
    but.add(self.timespan)
    sc = but.get_style_context()
    sc.add_class('ctrl-button')
    but.connect('clicked',self.open_numpad,self.timespan,{'min':0.0,'max':100000.0,'type':float,'polarity':False,'name':'Timespan (min)'})
    bx.pack_start(but,0,0,0)
    self.base_area.pack_start(bx,0,0,0)

    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.base_area.pack_start(divider,0,0,0)

    header_box = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
    lbl = Gtk.Label('Set Date / Time')
    self.add_style(lbl,["Label","font-22",'font-bold','text-black-color'])
    header_box.pack_start(lbl,1,1,1)
    #set current button
    but = Gtk.Button(width_request = 100)
    set_current = Gtk.Label()
    set_current.set_label('Set Current')
    self.add_style(set_current,['borderless-num-display','font-18','text-black-color'])
    but.add(set_current)
    sc = but.get_style_context()
    sc.add_class('ctrl-button')
    but.connect('clicked',self.load_current_time)
    header_box.pack_start(but,0,0,0)
    self.base_area.pack_start(header_box,1,1,1)

    self.hours = Gtk.SpinButton(orientation=Gtk.Orientation.VERTICAL,width_request = 80)
    self.hours.set_adjustment(Gtk.Adjustment(value=1, lower=0, upper=24, step_increment=1))
    self.hours.props.digits = 0
    self.add_style(self.hours,['font-36','text-black-color','spinbutton'])
    self.hours.connect('output', self.show_leading_zeros,1)
    hbox.pack_start(self.hours,0,0,0)

    lbl = Gtk.Label(':')
    self.add_style(lbl,['borderless-num-display','font-36','text-black-color','font-bold'])
    hbox.pack_start(lbl,0,0,0)
    
    self.minutes = Gtk.SpinButton(orientation=Gtk.Orientation.VERTICAL,width_request = 80)
    self.minutes.set_adjustment(Gtk.Adjustment(value=1, lower=0, upper=59,step_increment=1))
    self.minutes.props.digits = 0
    self.add_style(self.minutes,['font-36','text-black-color','spinbutton'])
    self.minutes.connect('output', self.show_leading_zeros,1)
    hbox.pack_start(self.minutes,0,0,0)

    lbl = Gtk.Label('.')
    self.add_style(lbl,['borderless-num-display','font-36','text-black-color','font-bold'])
    hbox.pack_start(lbl,0,0,0)

    self.seconds = Gtk.SpinButton(orientation=Gtk.Orientation.VERTICAL,width_request = 80)
    self.seconds.set_adjustment(Gtk.Adjustment(value=1, lower=0, upper=999,step_increment=1))
    self.seconds.props.digits = 0
    self.add_style(self.seconds,['font-36','text-black-color','spinbutton'])
    self.seconds.connect('output', self.show_leading_zeros,2)
    hbox.pack_start(self.seconds,0,0,0)

    #Blank Line
    lbl = Gtk.Label('')
    self.add_style(lbl,['borderless-num-display','font-18','text-black-color'])
    #hbox.pack_start(lbl,1,1,1)

    hbox.pack_start(vbox,0,0,0)

    bx = Gtk.Box(Gtk.Orientation.HORIZONTAL)
    self.add_style(bx,['padding'])
    lbl = Gtk.Label(label='Year ',width_request = 200)
    lbl.set_xalign(1.0)
    self.add_style(lbl,["Label","font-18",'font-bold','text-black-color'])
    bx.pack_start(lbl,0,0,0)
    self.year = Gtk.SpinButton(width_request = 120)
    self.year.set_orientation(Gtk.Orientation.HORIZONTAL)
    self.year.set_adjustment(Gtk.Adjustment(value=2022, lower=1900, upper=2050,step_increment=1))
    self.add_style(self.year,['font-18','text-black-color','spinbutton'])
    self.year.props.digits = 0
    bx.pack_start(self.year,0,0,0)
    vbox.pack_start(bx,0,0,0)

    bx = Gtk.Box(Gtk.Orientation.HORIZONTAL)
    self.add_style(bx,['padding'])
    lbl = Gtk.Label(label = 'Month ',width_request = 200)
    lbl.set_xalign(1.0)
    self.add_style(lbl,["Label","font-18",'font-bold','text-black-color'])
    bx.pack_start(lbl,1,1,1)
    months = Gtk.ListStore(int,str)
    lst = ['January','February','March','April','May','June','July','August','September','October','November','December']
    for num in range(len(lst)):
      months.append([num+1,lst[num]])
    self.months = Gtk.ComboBox.new_with_model(months)
    renderer_text = Gtk.CellRendererText()
    self.months.pack_start(renderer_text, True)
    self.months.add_attribute(renderer_text, "text", 1)
    self.months.connect("changed", self.on_month_combo_changed)
    self.add_style(self.months,['font-18','text-black-color'])
    bx.pack_start(self.months,0,0,0)
    vbox.pack_start(bx,0,0,0)

    bx = Gtk.Box(Gtk.Orientation.HORIZONTAL)
    self.add_style(bx,['padding'])
    lbl = Gtk.Label(label = 'Day ',width_request = 200)
    lbl.set_xalign(1.0)
    self.add_style(lbl,["Label","font-18",'font-bold','text-black-color'])
    bx.pack_start(lbl,1,1,1)
    self.day = Gtk.SpinButton(width_request = 120)
    self.day.set_orientation(Gtk.Orientation.HORIZONTAL)
    self.day.set_adjustment(Gtk.Adjustment(value=1, lower=1, upper=30,step_increment=1))
    self.day.props.digits = 0
    self.add_style(self.day,['font-18','text-black-color','spinbutton'])
    bx.pack_start(self.day,0,0,0)
    vbox.pack_start(bx,0,0,0)
    self.base_area.pack_start(hbox, 0, 0, 0)

    #Synchronize All Chart Times
    but = Gtk.Button(width_request = 100)
    self.sync_charts = Gtk.Label()
    self.sync_charts.set_label('Synchronize All Chart Times')
    but.connect("clicked", self.update_all_chart_times)
    self.add_style(self.sync_charts,['borderless-num-display','font-18','text-black-color'])
    but.add(self.sync_charts)
    sc = but.get_style_context()
    sc.add_class('ctrl-button')
    self.base_area.pack_start(but, 0, 0, 0)

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.save_settings,self.c_id,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    self.apply_button.connect('clicked',self.save_settings,self.c_id,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def load_current_time(self,*args):
    now = datetime.datetime.now()
    month = now.month
    d_t = {'hour':now.hour,'minute':now.minute,'second':now.second,'day':now.day,'month':month,'year':now.year}
    self.load_time(d_t)

  def load_time(self,d_t,*args):
    self.hours.set_value(d_t['hour'])
    self.minutes.set_value(d_t['minute'])
    self.seconds.set_value(d_t['second'])
    self.day.set_value(d_t['day'])
    self.months.set_active((d_t['month'])-1)
    self.year.set_value(d_t['year'])

  def update_all_chart_times(self,*args):
    for c_id in range(self.app.charts_number):
      self.save_settings('',c_id+1)

  def on_month_combo_changed(self, combo):
      tree_iter = combo.get_active_iter()
      if tree_iter is not None:
          model = combo.get_model()
          row_id, name = model[tree_iter][:2]
          #print("Selected: ID=%d, name=%s" % (row_id, name))
      else:
          entry = combo.get_child()

  def load_settings(self,*args):
    settings = self.db_session.query(self.Tbl).filter(self.Tbl.id == int(self.c_id)).first()
    if settings:
      self.timespan.set_text(str(settings.time_span))
      if settings.start_year == 1:
        self.load_current_time()
      else:
        self.hours.set_value(settings.start_hour)
        self.minutes.set_value(settings.start_minute)
        self.seconds.set_value(settings.start_second)
        self.day.set_value(settings.start_day)
        self.months.set_active((settings.start_month))
        self.year.set_value(settings.start_year)
    else:
      self.display_msg(msg="Chart Not Found")
      print("Chart Not Found")

  def display_msg(self,msg,*args):
    popup = PopupMessage(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      return True
    else:
      return False

  def save_settings(self,but,chart_id,auto_close,*args):
    settings = self.db_session.query(self.Tbl).filter(self.Tbl.id == int(chart_id)).first()
    if settings:
      settings.time_span =  int(self.timespan.get_label())
      settings.start_hour =  int(self.hours.get_value())
      settings.start_minute =  int(self.minutes.get_value())
      settings.start_second =  int(self.seconds.get_value())
      settings.start_year =  int(self.year.get_value())
      settings.start_month =  int(self.months.get_active())
      settings.start_day =  int(self.day.get_value())
      self.db_session.commit()
    self.app.charts[chart_id].reload_chart()
    if auto_close:
      self.close_popup()

  def show_leading_zeros(obj,spin_button,num,*args):
    adjustment = spin_button.get_adjustment()
    if num == 2:
      spin_button.set_text('{:03d}'.format(int(adjustment.get_value())))
    else:
      spin_button.set_text('{:02d}'.format(int(adjustment.get_value())))

  def add_style(self, item,style):
    sc = item.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def open_numpad(self,button,widget_obj,params,*args):
    numpad = ValueEnter(self,widget_obj,params)
    response = numpad.run()
    if response == Gtk.ResponseType.NO:
      pass
    else:
      pass
    numpad.destroy()

  def close_popup(self, *args):
    self.destroy()


class Export_ImportTagsPopup(Gtk.Dialog):
  def __init__(self, parent,app,export):
    Gtk.Dialog.__init__(self, '',None, Gtk.DialogFlags.MODAL,
                        (Gtk.STOCK_OK, Gtk.ResponseType.YES,
                        Gtk.STOCK_CANCEL, Gtk.ResponseType.NO)
                        )
    self.app = app
    self.parent = parent
    self.export = export
    self.connections_available = {}
    self.connect("response", self.on_response)
    self.result = {}
    self.get_available_connections()
    self.build_window()
    self.content_area = self.get_content_area()
    self.dialog_window = Gtk.Box(width_request=300,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window )
    ### - Title Bar- ###
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=300)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)

    ### - Base Area- ###
    self.base_area = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
    self.dialog_window.pack_start(self.base_area, 0, 0, 0)

    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=600)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)
    self.build_header()
    self.build_base()
    #self.build_footer()
    self.load_settings()
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(200, 200)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(False)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self,*args):
    #Save Button
    self.save_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Save.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.save_button.add(image)
    sc = self.save_button.get_style_context()
    sc.add_class('ctrl-button')
    bx = Gtk.Box()
    bx.pack_end(self.save_button,0,0,0)
    #self.title_bar.pack_start(bx,0,0,0)
    #self.save_button.connect('clicked',self.save_settings,self.c_id)

    #title
    if self.export:
      title = Gtk.Label(label='Export Tags')
    else:
      title = Gtk.Label(label='Import Tags')
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-22')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    #exit button
    self.exit_button = Gtk.Button(width_request = 20)
    self.exit_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.exit_button.add(image)
    #self.title_bar.pack_end(self.exit_button,0,0,1)
    sc = self.exit_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self,*args):

    #Label
    bx = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
    if self.export:
      lbl = Gtk.Label("Select Connection To Export Tags")
    else:
      lbl = Gtk.Label("Select Connection To Import Tags To")
    bx.pack_start(lbl,1,1,1)
    self.base_area.pack_start(bx,0,0,0)

    self.tag_sort = Gtk.ComboBoxText()
    self.tag_sort.set_entry_text_column(0)
    for x in self.connections_available:
      self.tag_sort.append_text(self.connections_available[x]['id'])
    self.tag_sort.set_active(1)
    sc = self.tag_sort.get_style_context()
    sc.add_class('ctrl-combo')
    self.base_area.pack_start(self.tag_sort,0,0,0)

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    #self.ok_button.connect('clicked',self.save_settings,self.c_id,True)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    #self.apply_button.connect('clicked',self.save_settings,self.c_id,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    #self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def load_settings(self,*args):
    pass

  def display_msg(self,msg,*args):
    popup = PopupMessage(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      return True
    else:
      return False

  def save_settings(self,but,chart_id,auto_close,*args):
    pass

  def add_style(self, item,style):
    sc = item.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def open_numpad(self,button,widget_obj,params,*args):
    numpad = ValueEnter(self,widget_obj,params)
    response = numpad.run()
    if response == Gtk.ResponseType.NO:
      pass
    else:
      pass
    numpad.destroy()

  def get_available_connections(self,*args):

    conx_items = ['id', 'connection_type', 'description']
    new_params = {}
    count = 1
    #self.connections_available = {0: {'id': '', 'connection_type': 0, 'description': ''}}
    self.connections_available = {}
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      for c in conx_items:
        new_params[c] = getattr(conx_obj, c)
      self.connections_available[count] = new_params
      new_params = {}
      count += 1

  def on_response(self, widget, response_id):
    self.result['conx_select'] = self.tag_sort.get_active_text()
  
  def get_result(self):
    return self.result

  def close_popup(self, *args):
    self.destroy()


class ImportUtility(Gtk.Dialog):
  #####################################Need to check if tag name has already been created
  #####################################Actually push data from each tag to database
  #####################################Handle boolean tags to be imported
  ######################### Verify comtrade import works

  def __init__(self, parent,app):
    super().__init__(transient_for = parent,flags=0)
    self.app = app
    ##$#self.conx_type = self.app.link.get('connection_types')
    self.conx_type = self.app.db.get('connection_types')
    self.db_session = self.app.settings_db.session
    self.db_model = self.app.settings_db.models['chart']
    self.Tbl = self.db_model
    self.data = []
    self.ANum = 0
    self.DNum = 0
    self.ch_Cfg = []
    self.NumSamp = 0
    self.start_time = 0.0
    self.end_time = 0.0
    self.big_box = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
    self.conversion = ['SEL','Comtrade']
    self.filterlist = ["*.CEV","*.cfg"]
    self.filter = {'SEL':'*.CEV'}
    self.new_conx = None
    
    self.build_window()
    self.content_area = self.get_content_area()
    self.dialog_window = Gtk.Box(height_request=800,width_request = 600,orientation=Gtk.Orientation.VERTICAL)
    self.content_area.add(self.dialog_window )
    ### - Title Bar- ###
    self.title_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=300)
    self.dialog_window.pack_start(self.title_bar,0,0,1)
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)

    ### - Base Area- ###
    self.base_area = Gtk.Box(spacing = 10,orientation=Gtk.Orientation.VERTICAL,margin = 20)
    self.scroll = Gtk.ScrolledWindow(width_request = 600,height_request = 600)
    self.scroll.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
    self.scroll.add(self.base_area)
    self.dialog_window.pack_start(self.scroll,1,1,1)

    ### -footer- ####
    divider = Gtk.Separator(orientation=Gtk.Orientation.HORIZONTAL)
    sc = divider.get_style_context()
    sc.add_class('Hdivider')
    self.dialog_window.pack_start(divider,0,0,1)
    self.footer_bar = Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL,height_request=20,width_request=800)
    self.dialog_window.pack_start(self.footer_bar,0,0,1)
    self.load_settings()
    self.build_header()
    self.build_base()
    self.build_footer()
    self.show_all()

  def build_window(self, *args):
    self.set_default_size(800, 800)
    self.set_decorated(False)
    self.set_border_width(10)
    self.set_keep_above(False)
    sc = self.get_style_context()
    sc.add_class("dialog-border")

  def build_header(self,*args):
    #Save Button
    self.save_button = Gtk.Button(width_request = 30)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Save.png', 30, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.save_button.add(image)
    sc = self.save_button.get_style_context()
    sc.add_class('ctrl-button')
    bx = Gtk.Box()
    bx.pack_end(self.save_button,0,0,0)
    #self.title_bar.pack_start(bx,0,0,0)
    #self.save_button.connect('clicked',self.save_settings,self.c_id)

    #title
    title = Gtk.Label(label='File Import Utility')
    sc = title.get_style_context()
    sc.add_class('text-black-color')
    sc.add_class('font-22')
    sc.add_class('font-bold')
    self.title_bar.pack_start(title,1,1,1)

    #exit button
    self.exit_button = Gtk.Button(width_request = 20)
    self.exit_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Close.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    self.exit_button.add(image)
    #self.title_bar.pack_end(self.exit_button,0,0,1)
    sc = self.exit_button.get_style_context()
    sc.add_class('exit-button')

  def build_base(self,*args):
    '''_________Body__________'''
    bbox1 = Gtk.Box(spacing = 10, orientation = Gtk.Orientation.HORIZONTAL)
    lbl = Gtk.Label('File Type')
    sc = lbl.get_style_context()
    sc.add_class('font-bold')
    bbox1.pack_start(lbl, True, True, 0)
    self.conv_type = Gtk.ComboBoxText()
    self.conv_type.set_size_request(100, 10)
    for item in self.conversion:
        self.conv_type.append_text(item)
    self.conv_type.set_active(0)
    self.conv_type.connect("changed", self.update_conv)
    bbox1.pack_start(self.conv_type,0,0,0)

    lbl = Gtk.Label('Import File')
    sc = lbl.get_style_context()
    sc.add_class('font-bold')
    bbox1.pack_start(lbl, True, True, 0)
    self.file_entry = Gtk.Label(width_request=300)
    self.file_entry.set_text('Select A File')
    sc = self.file_entry.get_style_context()
    sc.add_class('label-box')
    bbox1.pack_start(self.file_entry, False, False, 0)
    but = Gtk.Button(label = '...')
    but.set_property("width-request", 20)
    but.connect("clicked", self.fileChooser_open,None,self.filter)
    bbox1.pack_start(but, False, False, 0)
    self.base_area.pack_start(bbox1, False, False, 0)

    bbox2 = Gtk.Box(spacing = 10, orientation = Gtk.Orientation.HORIZONTAL)
    l = Gtk.Label('Export File')
    sc = l.get_style_context()
    sc.add_class('font-bold')
    bbox2.pack_start(l, True, True, 0)
    self.export_entry = Gtk.Entry(editable = False,width_request=300)
    self.export_entry.set_text('Select an Export Location')
    bbox2.pack_start(self.export_entry, False, False, 0)
    but = Gtk.Button(label = '...')
    but.set_property("width-request", 20)
    but.connect("clicked", self.fileChooser_save,None,self.filter)
    bbox2.pack_start(but, False, False, 0)
    #self.base_area.pack_start(bbox2, False, False, 0)
    self.base_content_area = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
    self.base_area.pack_start(self.base_content_area, True, True, 0)

    file_time = '0/0/0 - 00:00:00.0'
    header = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
    header.set_property("height-request", 40)
    lbl = Gtk.Label('File Time Stamp')
    self.add_style(lbl,["Label","font-18",'font-bold','text-black-color'])
    header.pack_start(lbl, 1, 1, 0)
    self.base_content_area.pack_start(header, 0, 0, 0)
    s = Gtk.Separator()
    self.base_content_area.pack_start(s, 0, 0, 0)
    box = Gtk.Box()
    box.set_homogeneous(True)
    box.set_halign(Gtk.Align.CENTER)
    box.set_spacing(10)
    box.set_property("height-request", 40)
    self.base_content_area.pack_start(box, 0, 0, 0)

    lbl = Gtk.Label('File Time:')
    self.add_style(lbl,["Label","font-14",'font-bold','text-black-color'])
    box.add(lbl)
    self.file_time = Gtk.Label(label=file_time)
    self.file_time.set_size_request(80, 25)
    box.add(self.file_time)
    lbl = Gtk.Label('Time Offset in (ms):')
    self.add_style(lbl,["Label","font-14",'font-bold','text-black-color'])
    box.add(lbl)
    self.time_offset = Gtk.Entry()
    self.time_offset.set_text('0')
    self.time_offset.set_tooltip_text("Enter A Time Correction Factor To The New File")
    self.time_offset.connect('changed', self.on_changed)
    box.add(self.time_offset)
    self.base_content_area.pack_start(Gtk.Separator(), 0, 0, 0)

    header2 = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
    header2.set_property("height-request", 40)
    lbl = Gtk.Label('Available Tags')
    self.add_style(lbl,["Label","font-18",'font-bold','text-black-color'])
    header2.pack_start(lbl, 1, 1, 0)
    self.base_content_area.pack_start(header2, 0, 0, 0)

  def build_footer(self):
    #CANCEL Button
    self.cancel_button = Gtk.Button(width_request = 100, height_request = 30)
    self.cancel_button.connect('clicked',self.close_popup)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Cancel')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.cancel_button.add(box)
    self.footer_bar.pack_end(self.cancel_button,0,0,1)
    sc = self.cancel_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #OK Button
    self.ok_button = Gtk.Button(width_request = 100, height_request = 30)
    self.ok_button.connect('clicked',self.convert)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('OK')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.ok_button.add(box)
    self.footer_bar.pack_end(self.ok_button,0,0,1)
    sc = self.ok_button.get_style_context()
    sc.add_class('ctrl-button-footer')

    #APPLY Button
    self.apply_button = Gtk.Button(width_request = 100, height_request = 30)
    #self.apply_button.connect('clicked',self.save_settings,self.c_id,False)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale('./ProcessPlot/Public/images/Return.png', 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    box = Gtk.Box()
    lbl = Gtk.Label('Apply')
    sc = lbl.get_style_context()
    sc.add_class('font-16')
    box.pack_start(lbl,1,1,1)
    #box.pack_start(image,0,0,0)
    self.apply_button.add(box)
    #self.footer_bar.pack_end(self.apply_button,0,0,1)
    sc = self.apply_button.get_style_context()
    sc.add_class('ctrl-button-footer')

  def load_settings(self,*args):
    pass

  def update_conv(self,*args):
      for item in range(len(self.conversion)):
          if self.conversion[item] == self.conv_type.get_active_text():
              self.filter[self.conversion[item]] = self.filterlist[item]

  def fileChooser_open(self,button,conx_sel,filt_pattern = ["*.xlsx"]):
      #Complete the import of the file, be sure to check if connection is running
      open_dialog = Gtk.FileChooserDialog(
          title="Please choose a file", parent=self, action=Gtk.FileChooserAction.OPEN
      )
      open_dialog.add_buttons(
          Gtk.STOCK_CANCEL,
          Gtk.ResponseType.CANCEL,
          Gtk.STOCK_OPEN,
          Gtk.ResponseType.OK,
      )
      filter = Gtk.FileFilter()
      #filter.set_name("*.xlsx")
      for pat in filt_pattern:
          filter.set_name(pat)
          filter.add_pattern(filt_pattern[pat])
      open_dialog.add_filter(filter)

      response = open_dialog.run()
      if response == Gtk.ResponseType.OK:
        self.fn = open_dialog.get_filename()
        if self.conv_type.get_active_text() == 'Comtrade':
            self.comtrade_import(self.fn)
        elif self.conv_type.get_active_text() == 'SEL':
            self.sel_import(self.fn)
      open_dialog.destroy()

  def fileChooser_save(self,button,conx_sel,filt_pattern = ["*.xlsx"]):
    save_dialog = Gtk.FileChooserDialog("Save As", self,
                                        Gtk.FileChooserAction.SAVE,
                                        (Gtk.STOCK_OK, Gtk.ResponseType.OK,
                                          Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL))
    save_dialog.set_current_name('File_Export')
    filter = Gtk.FileFilter()
    #filter.set_name("*.xlsx")
    for pat in filt_pattern:
        filter.set_name(pat)
        filter.add_pattern(filt_pattern[pat])
    #save_dialog.add_filter(filter)
    response = save_dialog.run()
    if response == Gtk.ResponseType.OK:
        self.foldname = save_dialog.get_filename()
        self.export_entry.set_text(self.foldname)
    else:
        # don't bother building the window
        self.destroy()
    save_dialog.destroy()

  def comtrade_import(self, *args):
      analog_dict = ['CNum', 'Name', 'Phase', 'Cir', 'Unit', 'Scale', 'Off', 'Skew', 'Min', 'Max','AD','Select']
      try:
          with open(self.fn,'r') as f:
              rows = f.readlines()
      except IOError as e:
          self.display_msg(msg='File Read Error: '+str(e))
      try:
          self.ANum = int((rows[1].split(',')[1]).replace('A', ''))      #Extract Number of Analogs from File
          self.DNum = int((rows[1].split(',')[2]).replace('D', ''))      #Extract Number of Digitals from File
          self.NumSamp = int((rows[4 + self.ANum + self.DNum].split(',')[1]).replace('/n', ''))
          for analog_chan in range(self.ANum):
              self.ch_Cfg.append(dict(zip(analog_dict, list(cell for cell in rows[2 + analog_chan].split(',')))))
              self.ch_Cfg[analog_chan]['AD'] = 'A'
              self.ch_Cfg[analog_chan]['Select'] = Gtk.CheckButton()
          for dig_chan in range(self.DNum):
              self.ch_Cfg.append(
                  dict(zip(analog_dict, list(cell for cell in rows[2 + self.ANum + dig_chan].split(',')))))
              self.ch_Cfg[dig_chan+self.ANum]['AD'] = 'D'
              self.ch_Cfg[dig_chan+self.ANum]['Select'] = Gtk.CheckButton()
          '''Start time'''
          day, month, year = str((rows[5 + self.ANum + self.DNum].split(',')[0])).split('/')
          hour, minute, temp = str((rows[5 + self.ANum + self.DNum].split(',')[1])).split(':')
          second, ms = temp.split('.')
          ms = '0.' + ms
          t = (int(year), int(month), int(day), int(hour), int(minute), int(second), 0, 0, 0)
          self.start_time = time.mktime(t) + float(ms)  # seconds after POSIX epoch
          '''Event time'''
          day, month, year = str((rows[6 + self.ANum + self.DNum].split(',')[0])).split('/')
          hour, minute, temp = str((rows[6 + self.ANum + self.DNum].split(',')[1])).split(':')
          second, ms = temp.split('.')
          ms = '0.' + ms
          t = (int(year), int(month), int(day), int(hour), int(minute), int(second), 0, 0, 0)
      except:
          self.display_msg(msg='File Header Corrupted')
          self.start_time = '0.0'
          self.NumSamp = 0

      try:
          with open(self.fn.replace('.cfg', '.dat'), 'r') as f:
              dfile = f.readlines()
          if len(dfile) != self.NumSamp:
              self.display_msg(msg='ComTradeObj: Data file samples do not match config. dat file may be corrupt. Using samples in data file')
              self.NumSamp = len(dfile)
          # time between samples is 1/Hz Sammple rate
          self.data.append(list((int(dfile[sample_num].split(',')[1]) * 0.000001) + self.start_time for sample_num in
                                range(self.NumSamp)))  # 2nd cell is microseconds after the start
          self.end_time = (self.data[0][-1])
          for item in range(len(self.ch_Cfg)):
              scaling = float(self.ch_Cfg[item]['Scale'])
              cell_num = 2 + item
              self.data.append(
                  list(float(dfile[sample_num].split(',')[cell_num]) * scaling for sample_num in range(self.NumSamp)))
          self.file_entry.set_text(self.fn)
          self.build_tree()
          #self.build_tree(self.fn,self.start_time,self.NumSamp,self.ch_Cfg)

      except IOError as e:
          self.display_msg(msg='Data File (.dat) Missing or not in cfg directory: ' + str(e))

  def sel_import(self, *args):
      analog_dict = ['CNum', 'Name', 'Phase', 'Cir', 'Unit', 'Scale', 'Off', 'Skew', 'Min', 'Max','AD','Select']
      try:
          with open(self.fn,'r') as f:
              rows = f.readlines()
      except IOError as e:
          self.display_msg(msg='File Read Error: ' + str(e))
      try:
        '''Start time'''
        #"MONTH","DAY","YEAR","HOUR","MIN","SEC","MSEC","0ACA"
        month, day, year, hour, minute, second, ms, na = str(rows[3]).split(',')
        t = (int(year), int(month), int(day), int(hour), int(minute), int(second), 0, 0, 0)
        # self.start_timestamp = time.mktime(t) + float(ms)  # seconds after POSIX epoch
        self.start_time = time.mktime(t) + float(ms)  # seconds after POSIX epoch
        '''Config Data'''
        #REC_NUM, REF_NUM, NUM_CH_A, NUM_CH_D, FREQ, NFREQ, SAM_CYC_A, SAM_CYC_D, NUM_OF_CYC, PRIM_VAL, CTR_IA, CTR_IB, CTR_IC, CTR_IN, CTR_IG, PTR_VA, PTR_VB, PTR_VC, PTR_VS, EVENT, LOCATION, GROUP, IA_A, IB_A, IC_A, IN_A, IG_A, VA_V, VB_V, VC_V, VG_V, VS_V, VDC_V, WDG_C, BRG_C, AMB_C, OTH_C, na = str(rows[5]).split(',')
        rec_names = []
        rec_names = str(rows[4]).split(',')    
        rec_num = []
        rec_num = str(rows[5]).split(',')
        rec_data = {}
        for x in range(len(rec_names)):
          rec_data[rec_names[x].replace('"','')] = rec_num[x]
        self.ANum = int(rec_data['NUM_CH_A'])
        self.DNum = int(rec_data['NUM_CH_D'])
        self.DNum = 0     #############################################NEED TO REMOVE IF YOU WANT DIGITALS TO BE IMPORTED
        NumSamp = (int(rec_data['SAM/CYC_A']) * int(rec_data['NUM_OF_CYC']))
        time_between_samples = (
            1.0 / (float(NumSamp - 16.0)))  # time between samples is total number of samples in 16 cycles
        '''Time between samples is 1/NumSamp'''
        time_data = []
        for num in range(NumSamp):
            time_data.append((float(num) * time_between_samples) + self.start_time)
        self.data.append(time_data)
        '''create list for analog channel names'''
        x=1
        for analog_chan in range(self.ANum):
            self.ch_Cfg.append(dict(zip(analog_dict,(str(x),(str(rows[6]).split(',')[analog_chan]).replace('"',''),'','','',1.0,'','','','','A' ))))
            self.ch_Cfg[analog_chan]['Select'] = False
            x+=1
        if self.DNum != 0:
            temp = (str(rows[6]).split(',')[-2])
        '''create list for digital channel names'''
        for dig_chan in range(self.DNum):
            #self.AnalNames.append(str(temp).split(' ')[item])
            self.ch_Cfg.append(dict(zip(analog_dict, (str(x), str(temp).split(' ')[dig_chan], '', '', '', 1.0, '', '', '','','D'))))
            self.ch_Cfg[dig_chan + self.ANum]['Select'] = False
            x += 1
        '''Bring Analog data into list of lists in self.data'''
        a_data = [[] for _ in range(self.ANum)]  # list of lists for all of the data +1 for the time
        for row in range(7, NumSamp + 7):
            for item in range(self.ANum):
                a_data[item].append(float(((str(rows[row]).split(',')[item]).replace('"', ''))))
        for item in range(self.ANum):
            self.data.append(a_data[item])
        d_data = []  # list of lists for all of the digital data, Contains all digital data as a large binary number
        for row in range(7, NumSamp + 7):
            d_data.append(((str(rows[row]).split(',')[self.ANum + 1]).replace("'", ' ').replace('"','')))
        '''convert hex to binary for digital status'''
        for row in range(len(d_data)):
            d_data[row] = bin(int(d_data[row], 16))[2:].zfill(len(d_data[row]) * 4)
            # section corrects for error in file when missing bytes
            if len(d_data[row]) != self.DNum:
                if row == 0:
                    d_data[row] = d_data[row + 1]
                else:
                    d_data[row] = d_data[row - 1]
        self.data.append(d_data)
        #Write data to temporary file
        with open(self.fn.replace('.CEV', '.dat'),'w') as f:
            temp = ''
            for row in range(len(self.data[0])):
                temp += str(row)+','
                for col in range(self.ANum+1):
                    temp += str(self.data[col][row])+','
                for z in range(self.DNum):
                    temp += str(self.data[self.ANum+1][row])[z] + ','
                rows = f.writelines(temp + "\n")
                temp = ''
        self.file_entry.set_text(self.fn)
        self.build_tree(self.fn.replace('.CEV', '.dat'), self.start_time, NumSamp, self.ch_Cfg)
      except IOError as e:
          self.display_msg(msg='File Import Error: '+str(e))

  def sel_import_old(self, *args):
      analog_dict = ['CNum', 'Name', 'Phase', 'Cir', 'Unit', 'Scale', 'Off', 'Skew', 'Min', 'Max','AD','Select']
      try:
          with open(self.fn,'r') as f:
              rows = f.readlines()
      except IOError as e:
          self.display_msg(msg='File Read Error: ' + str(e))

      '''Start time'''
      month, day, year, hour, minute, second, ms,na = str(rows[3]).split(',')
      t = (int(year), int(month), int(day), int(hour), int(minute), int(second), 0, 0, 0)
      # self.start_timestamp = time.mktime(t) + float(ms)  # seconds after POSIX epoch
      self.start_time = time.mktime(t) + float(ms)  # seconds after POSIX epoch
      '''Config Data'''
      rec_num, a_num, d_num, freqx, freqy, n_freq, acycles, samp_cycleD,numcycles,primvalue,na = str(rows[5]).split(',')
      self.ANum = int(a_num)
      self.DNum = int(d_num)
      NumSamp = (int(acycles) * int(numcycles))
      time_between_samples = (
          1.0 / (float(NumSamp - 16.0)))  # time between samples is total number of samples in 16 cycles
      '''Time between samples is 1/NumSamp'''
      time_data = []
      for num in range(NumSamp):
          time_data.append((float(num) * time_between_samples) + self.start_time)
      self.data.append(time_data)
      '''create list for analog channel names'''
      x=1
      for analog_chan in range(self.ANum):
          self.ch_Cfg.append(dict(zip(analog_dict,(str(x),(str(rows[6]).split(',')[analog_chan]).replace('"',''),'','','',1.0,'','','','','A' ))))
          self.ch_Cfg[analog_chan]['Select'] = False
          x+=1
      if self.DNum != 0:
          temp = (str(rows[6]).split(',')[-2])
      '''create list for digital channel names'''
      for dig_chan in range(self.DNum):
          #self.AnalNames.append(str(temp).split(' ')[item])
          self.ch_Cfg.append(dict(zip(analog_dict, (str(x), str(temp).split(' ')[dig_chan], '', '', '', 1.0, '', '', '','','D'))))
          self.ch_Cfg[dig_chan + self.ANum]['Select'] = False
          x += 1
      '''Bring Analog data into list of lists in self.data'''
      a_data = [[] for _ in range(self.ANum)]  # list of lists for all of the data +1 for the time
      for row in range(7, NumSamp + 7):
          for item in range(self.ANum):
              a_data[item].append(float(((str(rows[row]).split(',')[item]).replace('"', ''))))
      for item in range(self.ANum):
          self.data.append(a_data[item])
      d_data = []  # list of lists for all of the digital data, Contains all digital data as a large binary number
      for row in range(7, NumSamp + 7):
          d_data.append(((str(rows[row]).split(',')[self.ANum + 1]).replace("'", ' ').replace('"','')))
      '''convert hex to binary for digital status'''
      for row in range(len(d_data)):
          d_data[row] = bin(int(d_data[row], 16))[2:].zfill(len(d_data[row]) * 4)
          # section corrects for error in file when missing bytes
          if len(d_data[row]) != self.DNum:
              if row == 0:
                  d_data[row] = d_data[row + 1]
              else:
                  d_data[row] = d_data[row - 1]
      self.data.append(d_data)
      try:
          with open(self.fn.replace('.CEV', '.dat'),'w') as f:
              temp = ''
              for row in range(len(self.data[0])):
                  temp += str(row)+','
                  for col in range(self.ANum+1):
                      temp += str(self.data[col][row])+','
                  for z in range(self.DNum):
                      temp += str(self.data[self.ANum+1][row])[z] + ','
                  rows = f.writelines(temp + "\n")
                  temp = ''
          self.file_entry.set_text(self.fn)
          self.build_tree(self.fn.replace('.CEV', '.dat'), self.start_time, NumSamp, self.ch_Cfg)
      except IOError as e:
          self.display_msg(msg='File Write Error: '+str(e))

  def display_msg(self,msg,*args):
    popup = PopupMessage(self, msg=msg)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      return True
    else:
      return False

  def convert(self,*args):
    tags_selected_for_import = []
    duplicate_tag = []
    for row in self.liststore:
    # Print values of all columns
      if row[0]:  #searches liststore for all rows with a checkmark and adds names to a list
        if row[1] in tags_selected_for_import:  #Checks for duplicate tag names in import list
          duplicate_tag.append(row[1])
        else:
          tags_selected_for_import.append(row[1])
    if duplicate_tag:
      self.display_msg(msg='Duplicate Tag Names Exist Items Selected Only First Tag Will Be Imported')
    found = 0
    for item in range(len(self.ch_Cfg)):  #Figure out which tags were selected
        if self.ch_Cfg[item]['Name'] in tags_selected_for_import:
          self.ch_Cfg[item]['Select'] = True
          found += 1
    if found == 0:
        self.display_msg(msg='No Items Selected')
    else:
      self.add_connection_popup(None,None)
      self.close_popup()
      if self.new_conx == None:
        self.display_msg(msg='No Tags Imported')
      else:
        for tag in self.ch_Cfg:
          if tag['Select']:  #All self.ch_Cfg tags with select = True should be imported
            params = {"id": tag['Name'],
                                "connection_id": self.new_conx,
                                "description": '',
                                "datatype": 'FLOAT',
                                "address": tag['CNum'],
                    }
            self.create_tag(params)

  def add_connection_popup(self,button,bad_name,*args):
    ######################## Dont pass in params just generate default name for new local connection
    popup = AddConnectionPopup(self,None,self.app,['local'])
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      #results = (popup.get_result())
      #self.check_duplicate_name(results)
      return True
    else:
      return False

  def create_connection(self,params,*args):
    #should be passing in description and connection_type as a dictionary
    new_conx = self.app.link.new_connection({"id": params['id'],
                            "connection_type": params['connection_type'],
                            "description": params['description']
                            })
    ##$#conx_obj = self.app.link.get("connections").get(params['id'])
    conx_obj = self.app.db.get("connections").get(params['id'])
    if conx_obj != None:
      self.new_conx = params['id']
      self.app.link.save_connection(conx_obj)
      self.insert_connection_row(None,params)
      self.get_available_connections()
    else:
      #print('connection creation failed')
      self.display_msg(msg='Connection creation failed')

  def insert_connection_row(self,button,params,*args):
    pass

  def open_settings_popup(self,conx_id,*args):
    pass

  def get_available_connections(self,*args):
    conx_items = ['id', 'connection_type', 'description']
    new_params = {}
    self.connections_available = {}   #Clears out old list
    self.conx_obj_available = {}   #Clears out old list
    count = 0
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      for c in conx_items:
        new_params[c] = getattr(conx_obj, c)
      self.connections_available[count] = new_params
      self.conx_obj_available[conx_id] = conx_obj   #Hold reference to all conx objects
      new_params = {}
      count += 1

  def create_tag(self,params,*args):
    if 'address' not in params:
      params['address'] = '12'  # default address when not passed in because required to create tag
    ##$#conx_obj = self.app.link.get('connections').get(params['connection_id'])
    conx_obj = self.app.db.get('connections').get(params['connection_id'])
    conx_obj.new_tag({"id": params['id'],
                            "connection_id": params['connection_id'],
                            "description": params['description'],
                            "datatype": params['datatype'],
                            "address": params['address'],
    })
    tag_obj = conx_obj.get('tags').get(params['id'])
    if tag_obj != None:
      self.app.link.save_tag(tag_obj)
    params = self.get_tag_params(params['id'],params['connection_id'])
    self.insert_tag_row(None,params)

  def get_tag_params(self,tag_id,conx_id):
    new_params = {}
    ##$#conx_obj = self.app.link.get("connections").get(conx_id)
    conx_obj = self.app.db.get("connections").get(conx_id)
    if conx_obj != None:
      tag_items = conx_obj.return_tag_parameters()  #return list of tag parameters from the specific connection
      tag_obj = conx_obj.get('tags').get(tag_id)
      if tag_obj != None:
        for c in tag_items:
          new_params[c] = getattr(tag_obj, c)
        return new_params

  def insert_tag_row(self,button,params,*args):
    pass

  def add_style(self, item,style):
    sc = item.get_style_context()
    for sty in style:
      sc.add_class(sty)

  def build_tree(self, *args):
      try:
          temp = time.localtime(self.start_time)
          t_msec = (self.start_time) % 1
          t_msec = int(t_msec * 1000)
          file_time = str(temp.tm_mon) + '/' + str(temp.tm_mday) + '/' + str(
              temp.tm_year) + ' - ' + str(temp.tm_hour) + ':' + str(temp.tm_min) + ':' + str(temp.tm_sec) + '.' + str(
              t_msec).zfill(3)
      except:
          file_time = '0/0/0 - 00:00:00.0'
      self.file_time.set_text(file_time)

      #_____________________ Import list
      self.liststore = Gtk.ListStore(bool,str , str)
      self.treeview = Gtk.TreeView(self.liststore)
      #Watch for user clicks
      self.treeview.connect('button-press-event' , self.tree_item_clicked)
      self.treeview.set_rules_hint( True )
      self.add_style(self.treeview,['treeview'])

      #Add toggle button header
      connection_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/Tag.png', 25, 25)
      image = Gtk.Image(pixbuf=connection_icon)
      renderer_toggle = Gtk.CellRendererToggle()
      #renderer_toggle.set_property('cell-background','gray')
      self.tvcolumn_toggle = Gtk.TreeViewColumn('', renderer_toggle, active=0)
      h_but = self.tvcolumn_toggle.get_button() #Get reference to column header button
      c = h_but.get_child()
      c.add(image)  #add image to column header button
      #self.add_style(c,['treeview-header']) #add color to header
      c.show_all()

      #renderer_toggle.connect("toggled", self.tag_import_toggle)
      self.treeview.append_column(self.tvcolumn_toggle)
      self.tvcolumn_toggle.set_max_width(30)

      #Add Channel Name
      self.c_name = Gtk.CellRendererText()                         # create a CellRenderers to render the data\
      self.c_name.set_property("editable", True)
      self.c_name.set_property("xalign",0.5)
      #self.c_name.set_property("cell-background",'white')
      self.c_name.set_property("foreground",'black')
      self.c_name.connect('edited', self.name_check)
      col = Gtk.TreeViewColumn('Channel Name')
      c_but = col.get_button() #Get reference to column header button
      c = c_but.get_child()
      self.add_style(c,['treeview-header'])
      self.treeview.append_column(col)
      #col.set_expand(True)
      col.set_min_width(600)
      col.set_alignment(0.5)
      col.pack_start(self.c_name, True)
      col.set_sort_column_id(1)
      col.set_attributes(self.c_name,text=1)

      #Add Scaler
      self.scale = Gtk.CellRendererText()                         # create a CellRenderers to render the data\
      #self.scale.set_property("editable", True)                  #Makes text box editable
      self.scale.set_property("xalign",0.5)                        #Centers text in Box
      #self.scale.set_property("background",'white')              #changes box background to white
      self.scale.set_property("foreground",'black')
      #self.scale.connect('edited',self.change_num)
      self.scale_col = Gtk.TreeViewColumn('Scaler')
      s_but = self.scale_col.get_button() #Get reference to column header button
      c = s_but.get_child()
      self.add_style(c,['treeview-header'])
      self.treeview.append_column(self.scale_col)
      self.scale_col.pack_start(self.scale, False)
      self.scale_col.set_expand(False)
      self.scale_col.set_alignment(0.5)
      self.scale_col.set_sort_column_id(2)
      self.scale_col.set_attributes(self.scale,text=2)

      # make treeview searchable
      self.treeview.set_search_column(1)
      # Allow drag and drop reordering of rows
      self.treeview.set_reorderable(True)
      #Add treeview to base window
      scrolledwindow = Gtk.ScrolledWindow()
      scrolledwindow.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
      self.sc_box = Gtk.Box(orientation='vertical')
      self.sc_box.set_vexpand(True)
      scrolledwindow.add(self.sc_box)
      self.base_content_area.pack_start(scrolledwindow, 1, 1, 0)
      self.sc_box.add(self.treeview)

      self.add_import_rows()
      self.show_all()

  def add_import_rows(self,*args):
    for item in range(len(self.ch_Cfg)):
      self.liststore.append([ False,
                              self.ch_Cfg[item]['Name'],
                              str(self.ch_Cfg[item]['Scale']),
                                      ])
    self.show_all()

  def tree_item_clicked(self, treeview, event):
    pthinfo = treeview.get_path_at_pos(event.x, event.y)
    if event.button == 1: #left click
      if pthinfo != None:
        path,column,cellx,celly = pthinfo
        treeview.grab_focus()
        treeview.set_cursor(path,column,0)
        column_title = (column.get_title())
        #update currently active display
        selection = treeview.get_selection()
        tree_model, tree_iter = selection.get_selected()
        #If selected column is delete icon then initiate delete of connection
        if tree_iter != None:
          #gathers the Connection column name and connection type in the row clicked on
          t_id = tree_model[tree_iter][1]
          scale = tree_model[tree_iter][2]
          #checks if it is a toggle button click
          if column is self.tvcolumn_toggle:
            self.select_toggle('button',path,t_id)
          if column is self.scale_col:
            self.change_num(Gtk.CellRendererText(),path,scale)

      else:
        #unselect row in treeview
        selection = treeview.get_selection()
        selection.unselect_all()
    elif event.button == 3: #right click
      if pthinfo != None:
        path,col,cellx,celly = pthinfo
        treeview.grab_focus()
        treeview.set_cursor(path,col,0)
        rect = Gdk.Rectangle()
        rect.x = event.x
        rect.y = event.y + 10
        rect.width = rect.height = 1
        selection = treeview.get_selection()
        tree_model, tree_iter = selection.get_selected()
        popover = Gtk.Popover(width_request = 200)
        vbox = Gtk.Box(orientation=Gtk.Orientation.VERTICAL)
        if tree_iter is not None:
          #gathers the Tag name/Connection column text in the row clicked on
          t_id = tree_model[tree_iter][1]
          scale = tree_model[tree_iter][2]
          #popover to add display
          select_btn = Gtk.ModelButton(label="Select All", name=t_id)
          cb = lambda btn: self.select_all(t_id)
          select_btn.connect("clicked", cb)
          vbox.pack_start(select_btn, False, True, 10)
          clear_btn = Gtk.ModelButton(label="Clear All", name=t_id)
          cb = lambda btn:self.clear_all(t_id,tree_iter)
          clear_btn.connect("clicked", cb)
          vbox.pack_start(clear_btn, False, True, 10)
        popover.add(vbox)
        popover.set_position(Gtk.PositionType.RIGHT)
        popover.set_relative_to(treeview)
        popover.set_pointing_to(rect)
        popover.show_all()
        sc = popover.get_style_context()
        sc.add_class('popover-bg')
        sc.add_class('font-16')
        return
      else:
        return

  def select_toggle(self, widget, path,id):
    self.liststore[path][0] = not self.liststore[path][0]

  def on_changed(self, *args):
      text = self.time_offset.get_text().strip()
      self.time_offset.set_text(''.join([i for i in text if i in '0123456789']))

  def change_num(self, widget, path, text):
    #ref_val is the parameter used by the numpad to set the value in the treeview
    params = {'min':-100000.0,'max':100000.0,'type':float,'polarity':True,'name':'Scale','value':text,'ref_val':self.liststore[path]}
    self.open_numpad('',widget,params)

  def open_numpad(self,button,widget_obj,params,*args):
    numpad = ValueEnter(self,widget_obj,params)
    response = numpad.run()
    if response == Gtk.ResponseType.NO:
      pass
    else:
      pass
    numpad.destroy()

  def name_check(self, widget, path, text):
    temp = ''.join([i for i in text if i in 'aAbBcCdDeEfFgGhHiIjJkKlLmMnNoOpPqQrRsStTuUvVwWxXyYzZ1234567890_-()'])
    self.liststore[path][1] = temp

  def select_all(self, *args):
      for row in range(len(self.liststore)):
        path = Gtk.TreePath(row)
        self.liststore[path][0] = (True)

  def clear_all(self, *args):
      for row in range(len(self.liststore)):
        path = Gtk.TreePath(row)
        self.liststore[path][0] = (False)

  def close_popup(self, *args):
        self.destroy()


class Legend(object):
  ##################### collect all tags in the pen rows
  ##################### decide which tags to display have ability to disable in legend
  #####################
  ##################### 


  def __init__(self,app,legend_tab,parent,*args):
    self.app = app
    self.parent = parent    
    self.legend_tab = legend_tab
    self.connections_available = {}
    self.tags_available = {}
    self.pens_available = {}
    self.pen_column_names = ['id', 'chart_id', 'tag_id', 'connection_id', 'visible', 
                      'color', 'weight','scale_minimum','scale_maximum', 
                      'scale_lock', 'scale_auto']
    self.db_session = self.app.settings_db.session
    self.db_model = self.app.settings_db.models['pen']
    self.Tbl = self.db_model
    self.get_available_pens()
    self.get_available_connections()
    self.get_available_tags('')
    self.build_tree()
    self.update_pen_colors()

  def build_row(self,*args):

    #Display Status
    db_display_status = bool(True)
    box= Gtk.Box(orientation=Gtk.Orientation.HORIZONTAL)
    p_buf = GdkPixbuf.Pixbuf.new_from_file_at_scale(os.path.join(PUBLIC_DIR, 'images/Check.png'), 20, -1, True)
    image = Gtk.Image(pixbuf=p_buf)
    wid =CheckBoxWidget(30,30,image,db_display_status)
    self.display_status = wid.return_self()
    sc = self.display_status.get_style_context()
    sc.add_class('check-box')
    box.set_center_widget(self.display_status)
    #self.legend_tab .pack_start(box,1,1,1)

  def build_tree(self, *args):
      #_____________________ Import list
      def rendercolor(celllayout, cell, model, iter,*args):
        #if model.get_value(iter,3) == '#0000FF':
        db_color = str(model.get_value(iter,3))
        cell.set_property('background',db_color) 
          
      self.liststore = Gtk.ListStore(str, str, str, str)
      self.treeview = Gtk.TreeView(self.liststore)
      self.treeview.connect('button-press-event' , self.tree_item_clicked)             #Watch for user clicks
      sel = self.treeview.get_selection()                                               #Treeview not selectable
      sel.set_mode(Gtk.SelectionMode.NONE)                                               #Treeview not selectable
      self.treeview.set_rules_hint( True )
      self.treeview.set_grid_lines(Gtk.TreeViewGridLines.BOTH)

      #Add color label header
      color_icon = GdkPixbuf.Pixbuf.new_from_file_at_size('./ProcessPlot/Public/images/stop.png', 30, 30)
      image = Gtk.Image(pixbuf=color_icon)
      self.color_label = Gtk.CellRendererText()   
      self.color_label.set_property("editable", False)                
      col = Gtk.TreeViewColumn('', self.color_label)
      col.set_max_width(30)
      h_but = col.get_button() #Get reference to column header button
      c = h_but.get_child()
      c.add(image)
      c.show_all()  
      self.treeview.append_column(col)
      col.set_cell_data_func(self.color_label,rendercolor)

      #Add Tag Name
      self.c_name = Gtk.CellRendererText()                          # create a CellRenderers to render the data\
      self.c_name.set_property("editable", False)                    #Allows item to be clicked on and edited
      self.c_name.set_property("xalign",0.5)                        #Centers item 
      #self.c_name.set_property("cell-background",'white')          # Sets background color
      self.c_name.set_property("foreground",'black')
      col = Gtk.TreeViewColumn('Tag')
      c_but = col.get_button() #Get reference to column header button
      c = c_but.get_child()
      self.add_style(c,['treeview-header'])
      self.treeview.append_column(col)
      #col.set_expand(True)
      col.set_min_width(200)
      col.set_alignment(0.5)
      col.pack_start(self.c_name, True)
      col.set_sort_column_id(1)
      col.set_attributes(self.c_name,text=1)

      #Add Tag Value
      self.scale = Gtk.CellRendererText()                         # create a CellRenderers to render the data\
      self.scale.set_property("editable", False)                  #Makes text box editable
      self.scale.set_property("xalign",0.5)                        #Centers text in Box
      #self.scale.set_property("background",'white')              #changes box background to white
      self.scale.set_property("foreground",'black')
      #self.scale.connect('edited',self.change_num)
      self.scale_col = Gtk.TreeViewColumn('Value')
      s_but = self.scale_col.get_button() #Get reference to column header button
      c = s_but.get_child()
      self.add_style(c,['treeview-header'])
      self.treeview.append_column(self.scale_col)
      self.scale_col.pack_start(self.scale, False)
      self.scale_col.set_expand(False)
      self.scale_col.set_alignment(0.5)
      self.scale_col.set_max_width(50)
      self.scale_col.set_sort_column_id(2)
      self.scale_col.set_attributes(self.scale,text=2)

      #Add Color Number (Hidden)
      self.color_number = Gtk.CellRendererText()                          # create a CellRenderers to render the data\
      self.color_number.set_property("editable", False)                    #Allows item to be clicked on and edited
      self.color_number.set_property("xalign",0.5)                        #Centers item 
      self.color_number.set_property("visible",False)                     #Makes Colum Invisible
      col = Gtk.TreeViewColumn('')
      self.treeview.append_column(col)
      col.pack_start(self.color_number, True)
      col.set_sort_column_id(3)
      col.set_attributes(self.color_number,text=3)

      # make treeview searchable
      self.treeview.set_search_column(1)
      # Allow drag and drop reordering of rows
      self.treeview.set_reorderable(True)
      #Add treeview to base window
      scrolledwindow = Gtk.ScrolledWindow()
      scrolledwindow.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
      self.sc_box = Gtk.Box(orientation='vertical')
      self.sc_box.set_vexpand(True)
      scrolledwindow.add(self.sc_box)
      self.legend_tab.pack_start(scrolledwindow, 1, 1, 0)
      self.sc_box.add(self.treeview)
      self.add_rows()
      self.legend_tab.show_all()

  def add_rows(self,*args):
    for conx in self.tags_available:
      for tag_num in self.tags_available[conx]:
        color = self.get_pen_color(self.tags_available[conx][tag_num]['id'])
        self.liststore.append([ '',
                                str(self.tags_available[conx][tag_num]['id']),
                                str(self.tags_available[conx][tag_num]['value']),
                                color
                                        ])
    self.legend_tab.show_all()

  def update_pen_colors(self, *args):
    #Run this method if colors are updated with legend open
      for row in range(len(self.liststore)):
        path = Gtk.TreePath(row)
        color = self.get_pen_color(self.liststore[path][1])
        self.liststore[path][3]= color
      self.legend_tab.show_all()

  def tree_item_clicked(self, treeview, event):
    pthinfo = treeview.get_path_at_pos(event.x, event.y)
    if event.button == 1: #left click
      if pthinfo != None:
        path,column,cellx,celly = pthinfo
        treeview.grab_focus()
        treeview.set_cursor(path,column,0)
        column_title = (column.get_title())
        #update currently active display
        selection = treeview.get_selection()
        tree_model, tree_iter = selection.get_selected()
        self.open_pen_settings()

      else:
        #unselect row in treeview
        selection = treeview.get_selection()
        selection.unselect_all()
    elif event.button == 3: #right click
      pass

  def open_pen_settings(self, args=[]):
    popup = PenSettingsPopup(self.parent,self.app)
    response = popup.run()
    popup.destroy()
    if response == Gtk.ResponseType.YES:
      return True
    else:
      return False

  def open_numpad(self,button,widget_obj,params,*args):
    numpad = ValueEnter(self,widget_obj,params)
    response = numpad.run()
    if response == Gtk.ResponseType.NO:
      pass
    else:
      pass
    numpad.destroy()
  
  def add_style(self, item,style):
    sc = item.get_style_context()
    for items in sc.list_classes():
      #remove all default styles
      sc.remove_class(items)
    for sty in style:
      #add new styles
      sc.add_class(sty)

  def get_available_connections(self,*args):

    conx_items = ['id', 'connection_type', 'description']
    new_params = {}
    count = 1
    self.connections_available = {0: {'id': '', 'connection_type': 0, 'description': ''}}
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      for c in conx_items:
        new_params[c] = getattr(conx_obj, c)
      self.connections_available[count] = new_params
      new_params = {}
      count += 1

  def get_available_tags(self,c_id,*args):
    self.tags_available = {}
    new_params = {}
    count = 1
    self.conx_tags = {}
    ##$#for conx_id,conx_obj in self.app.link.get('connections').items():
    for conx_id,conx_obj in self.app.db.get('connections').items():
      if conx_obj:
        tag_items = conx_obj.return_tag_parameters()  #return list of tag parameters from the specific connection
        for tag_id,tag_obj in conx_obj.get('tags').items():
          for c in tag_items:
            new_params[c] = getattr(tag_obj, c)
          self.conx_tags[count] = new_params
          new_params = {}
          count += 1
        self.tags_available[conx_id]= self.conx_tags
        self.conx_tags = {}
        count = 1

  def get_available_pens(self,*args):
    params = {}
    pen = {}
    settings = self.db_session.query(self.Tbl).order_by(self.Tbl.id)
    for pen_param in settings:
      for c in self.pen_column_names:
        pen[c] = getattr(pen_param, c)
      params[pen['id']] = pen
      pen = {}
    self.pens_available = params

  def get_pen_color(self,tag_id,*args):
    entry = self.db_session.query(self.Tbl).filter(self.Tbl.tag_id == tag_id).first()
    if entry:
      return entry.color
    else:
      return '#b1b1b1'  #color gray

